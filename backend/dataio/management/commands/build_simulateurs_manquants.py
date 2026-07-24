"""
Construit les classeurs simulateurs absents du dossier institutionnel
(`CAFE_ARABICA`, `MANIOC`, `RIZ` — dette CLAUDE.md §6).

    python manage.py build_simulateurs_manquants            # écrit les manquants
    python manage.py build_simulateurs_manquants --force    # réécrit ceux qui existent
    python manage.py build_simulateurs_manquants --check    # ne rien écrire, seulement dire

Puis, comme pour n'importe quel autre classeur — c'est le point :

    python manage.py ingest_simulateurs

POURQUOI UNE COMMANDE, ET AUCUNE MIGRATION
------------------------------------------
Règle d'or de `dataio` : un nouveau fichier n'entraîne JAMAIS de migration. Un
classeur est de la donnée — chaque feuille devient une `DataTable`, chaque
en-tête une `DataColumn`, chaque ligne un `DataRecord`. Les trois simulateurs
créés ici empruntent donc exactement le chemin des quatorze existants : dépôt
dans le dossier standard, `ingest_simulateurs` (inspect → commit), lecture par
`credits.referentiel_loader.charger_depuis_simulateur`. Pas une ligne de schéma,
pas un `if` de plus dans le moteur.

D'OÙ VIENNENT LES CHIFFRES — ET CEUX QU'ON REFUSE D'ÉCRIRE
----------------------------------------------------------
Principe 1 : ce qui sert à scorer vient de la base, jamais d'une estimation. Le
seul référentiel qu'AGRICAP possède déjà pour ces trois filières est le
`reference_data.ValueChain` ACTIF — coût par hectare et répartition par module,
validés en maker-checker. C'est lui, et lui seul, qui alimente les feuilles 4 et
5 : le total est le coût/ha de l'institution, la ventilation est sa propre
grille de poids. Rien n'est réparti « à la main » — c'est précisément l'erreur
que `referentiel_loader` documente (850 USD de semences inventés là où le
classeur maïs en portait 126,60).

Ce que le référentiel ne porte PAS reste VIDE, et le classeur le dit :
  - rendement attendu et prix prévu (feuilles 3 et 8) — aucune table AGRICAP ne
    donne un rendement pour ces filières ; les inventer fabriquerait un DSCR ;
  - plages de vraisemblance (feuille 18) — même raison ;
  - valeurs de garanties (feuille 15) — elles appartiennent au dossier, pas au
    référentiel.

Conséquence assumée et vérifiée par les tests : `charger_depuis_simulateur`
construit `couts_modules` (la « fiabilité technique », 25 % de la note, était
déclarée NON CALCULABLE pour ces trois filières) et laisse `rendement_ref` vide.
Un référentiel partiellement inventé est pire qu'un référentiel absent, parce
qu'il score sans le dire.

LA FORME VIENT DU CLASSEUR MODÈLE, PAS D'ICI
--------------------------------------------
Les dix-neuf feuilles, leurs en-têtes, leurs formules et leur ordre ne sont pas
retapés dans ce fichier : on part du classeur institutionnel de la MÊME famille
(`AGRICAP_FIN_SIM_01_Cereales_Mais.xlsx` pour le riz, etc.) et on n'écrit que
les cellules de saisie. Retaper la structure en Python en ferait une seconde
source de vérité, qui divergerait au premier classeur révisé — et
`dataio.services.detect_kind` comme `_find_header_row` dépendent de cette forme
exacte.
"""
from __future__ import annotations

import os
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from reference_data.models import MODULE_WEIGHT_KEYS, ReferenceFileUpload, ValueChain

#: Sous-dossier standard des simulateurs par chaîne — le même que celui que lit
#: `credits.management.commands.ingest_simulateurs` (une seule adresse).
DEFAULT_SUBDIR = "Agricap FIN simulateur Par Chaine"

#: Superficie de référence du classeur, en hectares. `1` n'est pas un choix
#: arbitraire : c'est l'unité DU référentiel (`cost_per_hectare_usd`). Le loader
#: divise les totaux du cycle par cette dimension ; à 1 ha, le coût de référence
#: par module ressort exactement égal à la grille de l'institution, sans arrondi
#: intermédiaire ni facteur d'échelle à justifier.
SUPERFICIE_REFERENCE = 1

#: Les trois filières actives sans classeur (CLAUDE.md §6). `numero` est le code
#: de CHAÎNE du référentiel v3 (`referentiel.chains`, « 01 »–« 14 »), pas un
#: numéro de culture : le riz est une céréale (01), le manioc un tubercule (03),
#: le café relève de l'agroforesterie (12). Créer un « 15 » inventerait une
#: quinzième chaîne et rouvrirait la dette des nomenclatures parallèles
#: (principe 6). `modele` est le classeur de la même famille dont on reprend la
#: forme.
FILIERES_MANQUANTES: dict[str, dict[str, str]] = {
    "RIZ": {
        "numero": "01",
        "famille": "Cereales",
        "culture": "Riz",
        "modele": "AGRICAP_FIN_SIM_01_Cereales_Mais.xlsx",
    },
    "MANIOC": {
        "numero": "03",
        "famille": "Tubercules",
        "culture": "Manioc",
        "modele": "AGRICAP_FIN_SIM_03_Tubercules_PatateDouce.xlsx",
    },
    # `Cafe_Arabica` et non `CafeArabica` : `credits.dataio_simulator._find_source`
    # apparie la filière au classeur par MOTS du nom de fichier. Collé, le nom ne
    # partagerait aucun mot avec le code `CAFE_ARABICA` et le simulateur ne serait
    # jamais retenu pour sa propre filière. Le motif de nommage l'autorise
    # (`referentiel_loader._NOM_SIMULATEUR` : la culture accepte l'underscore).
    "CAFE_ARABICA": {
        "numero": "12",
        "famille": "Agroforesterie",
        "culture": "Cafe_Arabica",
        "modele": "AGRICAP_FIN_SIM_12_Agroforesterie_TaungyaAcaciaMais.xlsx",
    },
}

BANDEAU_ACCUEIL = (
    "CLASSEUR DE RÉFÉRENCE INSTITUTIONNEL — établi pour {superficie} hectare depuis le "
    "référentiel ValueChain actif « {code} » (coût/ha et poids par module validés en "
    "maker-checker). Les onglets 4 et 5 portent donc des coûts de RÉFÉRENCE, pas le plan "
    "d'un demandeur."
)

BANDEAU_BESOINS = (
    "Une ligne par rubrique = coût de référence de la rubrique pour {superficie} hectare. "
    "Le détail par poste (quantités, prix unitaires) n'est pas renseigné : aucun "
    "référentiel AGRICAP ne le porte pour cette filière, et l'inventer fabriquerait des "
    "écarts techniques faux."
)

BANDEAU_VENTES = (
    "RENDEMENT ET PRIX NON RENSEIGNÉS : aucun référentiel AGRICAP ne porte de rendement "
    "ni de prix pour cette filière. Tant qu'ils ne sont pas saisis ici, les onglets 9 à "
    "13 et 17 (EBE, DSCR, stress, score) ne veulent rien dire — ils calculent sur des "
    "recettes nulles. Le classeur préfère le dire plutôt que de proposer un chiffre."
)


def _q2(value: Decimal) -> Decimal:
    return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _nom_fichier(spec: dict[str, str]) -> str:
    return f"AGRICAP_FIN_SIM_{spec['numero']}_{spec['famille']}_{spec['culture']}.xlsx"


def repartir_couts(chain: ValueChain, superficie: int = SUPERFICIE_REFERENCE) -> dict[str, Decimal]:
    """Coût du cycle par module = poids × coût/ha × superficie, en `Decimal`.

    Le dernier module absorbe l'écart d'arrondi pour que la somme des rubriques
    soit RIGOUREUSEMENT le coût total : la feuille 5 doit égaler la feuille 4 au
    centime (principe 5), et un total qui « tombe à peu près » se paierait en
    contrôle de cohérence rouge sur chaque dossier de la filière.
    """
    total = _q2(Decimal(chain.cost_per_hectare_usd) * superficie)
    poids = chain.module_weights or {}
    montants: dict[str, Decimal] = {}
    cumul = Decimal("0.00")
    for module in MODULE_WEIGHT_KEYS:
        part = Decimal(str(poids.get(module, 0)))
        montant = _q2(total * part / Decimal(100))
        montants[module] = montant
        cumul += montant
    ecart = total - cumul
    if ecart and montants:
        dernier = MODULE_WEIGHT_KEYS[-1]
        montants[dernier] = _q2(montants[dernier] + ecart)
    return montants


def _f(value: Decimal) -> float:
    """`Decimal` → nombre écrit dans la cellule.

    Excel ne stocke que des flottants : la conversion a lieu ici, à l'écriture,
    et jamais dans un calcul. Toute la répartition ci-dessus s'est faite en
    `Decimal` (principe 4) ; ce `float` est un format de sortie, pas une étape
    de calcul.
    """
    return float(value)


def construire_classeur(chain: ValueChain, spec: dict[str, str], modele_path: str,
                        superficie: int = SUPERFICIE_REFERENCE):
    """Classeur openpyxl du simulateur de `chain`, calqué sur `modele_path`.

    N'écrit QUE des cellules de saisie ; formules, en-têtes, ordre des feuilles
    et mise en forme viennent du modèle.
    """
    wb = openpyxl.load_workbook(modele_path, data_only=False)
    # Le classeur sort d'openpyxl sans valeurs calculées : on demande à Excel de
    # tout recalculer à l'ouverture plutôt que d'afficher des cellules vides.
    wb.calculation.fullCalcOnLoad = True

    montants = repartir_couts(chain, superficie)
    total = sum(montants.values(), Decimal("0.00"))
    taux_ratio = float(Decimal(chain.base_rate) / Decimal(100))

    # ── 1_Accueil_Parametres — identité et paramètres de prêt ────────────────
    ws = wb["1_Accueil_Parametres"]
    ws["A2"] = BANDEAU_ACCUEIL.format(superficie=superficie, code=chain.code)
    ws["B7"] = f"Référence {chain.label} — {superficie} hectare"
    ws["B8"] = f"REF-{chain.code}"
    ws["B9"] = "AGRICAP — référentiel institutionnel"
    ws["B10"] = chain.label
    ws["B11"] = None                       # zone : propre au dossier, pas au référentiel
    ws["B12"] = None                       # campagne : idem
    ws["B15"] = chain.cycle_months
    ws["B16"] = taux_ratio
    ws["B17"] = chain.cycle_months         # durée de crédit alignée sur le cycle
    ws["B19"] = None                       # analyste : propre au dossier
    # Différé : le référentiel ne le porte pas. `0` est le comportement neutre
    # (mensualités dès le mois 1) — un différé « par défaut » de 5 mois hérité du
    # modèle maïs serait une hypothèse de trésorerie prise à la place du comité.
    ws["B20"] = 0

    # ── 2_Identification_Projet — la DIMENSION que lit le loader ─────────────
    ws = wb["2_Identification_Projet"]
    ws["B5"] = None
    ws["B6"] = None
    ws["B7"] = chain.label                 # « Production principale » → nom de filière
    ws["B8"] = superficie                  # « Superficie exploitée (ha) » → dimension
    ws["B9"] = None
    ws["B10"] = None
    ws["B12"] = "; ".join(chain.eligible_guarantees or []) or None
    ws["B13"] = 0                          # apport propre : propre au dossier

    # ── 3_Parametres_Techniques — calendrier et rendement : NON renseignés ───
    ws = wb["3_Parametres_Techniques"]
    for row in range(5, 11):
        for col in ("B", "C", "E", "F", "G", "H", "I", "J"):
            ws[f"{col}{row}"] = None
    ws["J5"] = ("Calendrier cultural et rendement à compléter — le référentiel "
                "ValueChain ne porte ni phases ni rendement pour cette filière.")

    # ── 4_Besoins_Financiers — une ligne par rubrique, valeurs de référence ──
    from credits.needs_sheet import MODULE_LABELS   # nomenclature unique (principe 6)

    ws = wb["4_Besoins_Financiers"]
    ws["A3"] = BANDEAU_BESOINS.format(superficie=superficie)
    premiere, derniere = 5, 35             # lignes de données du modèle
    for offset, module in enumerate(MODULE_WEIGHT_KEYS):
        row = premiere + offset
        rubrique = MODULE_LABELS[module]
        montant = montants[module]
        ws[f"A{row}"] = offset + 1
        ws[f"B{row}"] = rubrique
        ws[f"C{row}"] = f"Coût de référence — {rubrique.lower()}"
        ws[f"D{row}"] = "ha"
        ws[f"E{row}"] = superficie
        ws[f"F{row}"] = _f(_q2(montant / superficie))
        ws[f"G{row}"] = 1
        # Valeur et non formule : la feuille est un CONSTAT de référence, et c'est
        # cette valeur que `dataio` relit (`load_workbook(data_only=True)`), qui ne
        # voit d'une formule que sa valeur en cache — absente d'un fichier
        # qu'Excel n'a jamais ouvert.
        ws[f"H{row}"] = _f(montant)
        ws[f"I{row}"] = "Tout le cycle"
        ws[f"J{row}"] = "Crédit"
        ws[f"K{row}"] = f"Poids référentiel : {chain.module_weights.get(module, 0)} %"
        ws[f"L{row}"] = None
    for row in range(premiere + len(MODULE_WEIGHT_KEYS), derniere + 1):
        for col in "ABCDEFGHIJKL":
            ws[f"{col}{row}"] = None
    ws[f"H{derniere + 1}"] = _f(total)     # ligne TOTAL du modèle

    # ── 5_Synthese_Besoins — ce que lit `charger_depuis_simulateur` ──────────
    # Les LIBELLÉS de rubrique sont réécrits, pas hérités du modèle : certains
    # classeurs institutionnels renomment localement une rubrique (le classeur 12
    # appelle l'équipement « Plants & investissements agroforestiers »), et
    # `credits.needs_sheet.rubrique_to_module` ne la reconnaît alors plus — le
    # module disparaît du référentiel sans un mot. On écrit la nomenclature
    # canonique (principe 6), seule garantie que les huit modules soient relus.
    ws = wb["5_Synthese_Besoins"]
    for offset, module in enumerate(MODULE_WEIGHT_KEYS):
        row = 5 + offset
        ws[f"A{row}"] = MODULE_LABELS[module]
        ws[f"B{row}"] = _f(montants[module])
        ws[f"C{row}"] = None
        ws[f"D{row}"] = f"{chain.module_weights.get(module, 0)} % du coût de référence"
    ws["A13"] = "TOTAL GENERAL"
    ws["B13"] = _f(total)

    # ── 8_Previsions_Ventes — volontairement vide, et qui le dit ─────────────
    ws = wb["8_Previsions_Ventes"]
    ws["B2"] = None                        # rendement attendu
    ws["A5"] = chain.label
    ws["D5"] = None                        # taux de perte
    ws["F5"] = None                        # prix unitaire
    for col in "ABCDEFGHIJ":               # ligne « sous-produit » du modèle
        ws[f"{col}6"] = None
    ws["A11"] = BANDEAU_VENTES

    # ── 14_Annexes_Hypotheses — hypothèses non tenues, sauf le taux ──────────
    ws = wb["14_Annexes_Hypotheses"]
    for row in (5, 6, 7):
        ws[f"B{row}"] = None
        ws[f"C{row}"] = None
        ws[f"E{row}"] = "À documenter par l'analyste — non porté par le référentiel"
    ws["C8"] = taux_ratio
    ws["B8"] = f"Taux de base {chain.code}"
    ws["E8"] = "Repris de ValueChain.base_rate (référentiel actif)"
    ws["C9"] = f"ValueChain « {chain.code} »"
    ws["E9"] = "Coûts par module : poids du référentiel × coût/ha"

    # ── 15_Garanties_Collateraux — appartiennent au dossier, pas au modèle ───
    ws = wb["15_Garanties_Collateraux"]
    for row in range(5, 10):
        ws[f"C{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None

    # ── 18_Controles_Vraisemblance — pas de plage inventée ───────────────────
    ws = wb["18_Controles_Vraisemblance"]
    ws["A7"] = "Densité de plantation (unités/ha)"
    for row in range(5, 11):
        ws[f"C{row}"] = None               # Réf. min
        ws[f"D{row}"] = None               # Réf. max
        ws[f"F{row}"] = None               # commentaire de référence
    ws["F5"] = ("Plages à établir par la boucle d'apprentissage (principe 10) — "
                "aucune référence indicative disponible pour cette filière.")

    # ── 19_Suivi_Post_Decaissement — jalons rendus agnostiques de la culture ─
    ws = wb["19_Suivi_Post_Decaissement"]
    ws["E6"] = "Implantation conforme au plan (dimension, densité, état sanitaire)"
    ws["E8"] = "Estimation de production au champ, organisation de la récolte"

    return wb


class Command(BaseCommand):
    help = ("Construit les classeurs simulateurs manquants (CAFE_ARABICA, MANIOC, RIZ) "
            "depuis le référentiel ValueChain actif, sans migration ni schéma nouveau.")

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", default=None,
            help=f"Dossier de sortie (défaut : DOCUMENT_EXCEL_DIR/« {DEFAULT_SUBDIR} »).",
        )
        parser.add_argument(
            "--force", action="store_true",
            help="Réécrit un classeur déjà présent (par défaut il est conservé).",
        )
        parser.add_argument(
            "--check", action="store_true",
            help="N'écrit rien : dit seulement ce qui manque et ce qui bloquerait.",
        )

    def _dire(self, message: str, style=None) -> None:
        """Rapport console, muet à `--verbosity 0` (la commande tourne en test)."""
        if self._verbosity:
            self.stdout.write(style(message) if style else message)

    def handle(self, *args, **opts):
        self._verbosity = opts.get("verbosity", 1)
        directory = opts["output"] or os.path.join(settings.DOCUMENT_EXCEL_DIR, DEFAULT_SUBDIR)
        if not os.path.isdir(directory):
            raise CommandError(f"Dossier introuvable : {directory}")

        actives = {
            vc.code: vc
            for vc in ValueChain.objects.filter(
                source_file__status=ReferenceFileUpload.Status.ACTIVE, active=True,
            )
        }
        self._dire(f"Dossier : {directory}")
        self._dire(f"Référentiel actif : {len(actives)} filière(s).")

        ecrits = sautes = bloques = 0
        for code, spec in FILIERES_MANQUANTES.items():
            nom = _nom_fichier(spec)
            cible = os.path.join(directory, nom)
            chain = actives.get(code)

            if chain is None:
                # On ne fabrique pas un classeur sans référentiel : ses coûts
                # seraient inventés, ce que la commande existe pour éviter.
                bloques += 1
                self._dire(
                    f"  [!]  {nom} : « {code} » absent du referentiel ACTIF "
                    f"(reference_data.ValueChain) - rien n'est ecrit, aucun cout "
                    f"n'est devine.", self.style.WARNING)
                continue

            if os.path.exists(cible) and not opts["force"]:
                sautes += 1
                self._dire(f"  =  {nom} : déjà présent, conservé (--force pour réécrire).")
                continue

            modele = os.path.join(directory, spec["modele"])
            if not os.path.isfile(modele):
                bloques += 1
                self._dire(
                    f"  [!]  {nom} : classeur modèle « {spec['modele']} » introuvable - "
                    f"la forme d'un simulateur n'est pas retapee ici.", self.style.WARNING)
                continue

            montants = repartir_couts(chain)
            total = sum(montants.values(), Decimal("0.00"))
            if opts["check"]:
                self._dire(
                    f"  ?  {nom} : à écrire — {total} USD/ha répartis sur "
                    f"{len([m for m in montants.values() if m])} module(s).")
                continue

            wb = construire_classeur(chain, spec, modele)
            wb.save(cible)
            wb.close()
            ecrits += 1
            self._dire(
                f"  +  {nom} : ecrit - {total} USD pour {SUPERFICIE_REFERENCE} ha "
                f"(coût/ha référentiel : {chain.cost_per_hectare_usd}).", self.style.SUCCESS)

        style = self.style.SUCCESS if bloques == 0 else self.style.WARNING
        self._dire(
            f"{ecrits} classeur(s) ecrit(s), {sautes} conserve(s), {bloques} bloque(s). "
            f"Ingestion : python manage.py ingest_simulateurs", style)
