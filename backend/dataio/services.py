"""
Services de la couche générique : inspecter (aperçu) et enregistrer (commit manuel).

- `inspect(source)` : lit le classeur, détecte le type, calcule un aperçu (feuilles,
  en-têtes, échantillon, nb de lignes) SANS écrire de lignes. Appelé à l'upload.
- `commit(source)` : écrit DataTable/DataColumn/DataRecord pour chaque feuille, gère
  l'historique (nouvelle révision courante, anciennes conservées), et alimente les
  tables typées si c'est un référentiel (hybride).
"""
from __future__ import annotations

import unicodedata

import openpyxl
from django.db import transaction
from django.utils import timezone

from referentiel.chains import BY_SHEET
from referentiel.ingest import (
    CALIBRATION_SHEET, ingest_workbook,
    rebuild_chain_from_records, rebuild_config_from_records,
)
from referentiel.models import ReferentielVersion
from referentiel.range_parser import to_number, to_range
from .models import (
    DataColumn, DataRecord, DataSource, DataTable,
    KIND_ANNEXE, KIND_AUTRE, KIND_REFERENTIEL, KIND_SIMULATEUR,
    STATUS_COMMITTED,
)

SIMULATEUR_MARKERS = {"1_Accueil_Parametres", "4_Besoins_Financiers", "8_Previsions_Ventes"}
MAX_PREVIEW_ROWS = 8


def _norm(text) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFD", str(text))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


def dataset_key(name: str) -> str:
    """Clé logique d'un dataset (nom de fichier normalisé) — base du versionnage."""
    return _norm(name)


def detect_kind(sheetnames: list[str]) -> str:
    
    sheets = set(sheetnames)
    if any(BY_SHEET.get(s) for s in sheets) and len(sheets & set(BY_SHEET)) >= 5:
        return KIND_REFERENTIEL
    if len(sheets & SIMULATEUR_MARKERS) >= 2 and len(sheets) >= 15:
        return KIND_SIMULATEUR
    if "4_Besoins_Financiers" in sheets:
        return KIND_ANNEXE
    return KIND_AUTRE


def _infer_dtype(values: list) -> str:
    """Type inféré d'une colonne à partir d'un échantillon de valeurs."""
    seen = [v for v in values if v not in (None, "")]
    if not seen:
        return DataColumn.TEXT
    n_num = n_pct = n_rng = 0
    for v in seen:
        s = str(v)
        if "%" in s:
            n_pct += 1
            continue
        lo, hi = to_range(v)
        if lo is not None and hi is not None and lo != hi:
            n_rng += 1
        elif isinstance(v, (int, float)) or (to_number(v) is not None and not any(c.isalpha() for c in s)):
            n_num += 1
    total = len(seen)
    if n_pct / total > 0.5:
        return DataColumn.PERCENT
    if n_rng / total > 0.5:
        return DataColumn.RANGE
    if n_num / total > 0.6:
        return DataColumn.NUMBER
    return DataColumn.TEXT


def _find_header_row(rows: list[tuple], limit: int = 8) -> int:
    """Ligne d'en-tête = 1re ligne (0-based) avec ≥3 cellules non vides et courtes."""
    for i, row in enumerate(rows[:limit]):
        cells = [c for c in row if c not in (None, "")]
        short = [c for c in cells if len(str(c)) <= 40]
        if len(cells) >= 3 and len(short) >= max(2, len(cells) // 2):
            return i
    return 0


def _sheet_matrix(ws, max_rows: int | None = None) -> list[tuple]:
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        out.append(row)
        if max_rows and i + 1 >= max_rows:
            break
    return out


def inspect(source: DataSource) -> dict:
    """Calcule l'aperçu (sans écrire de lignes) et le stocke sur la source."""
    wb = openpyxl.load_workbook(source.file.path, data_only=True, read_only=True)
    kind = detect_kind(wb.sheetnames)
    tables_preview = []
    for pos, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows = _sheet_matrix(ws, max_rows=MAX_PREVIEW_ROWS + 12)
        hr = _find_header_row(rows)
        header = [str(c).strip() if c is not None else "" for c in (rows[hr] if hr < len(rows) else ())]
        data_rows = rows[hr + 1:]
        sample = []
        for r in data_rows[:MAX_PREVIEW_ROWS]:
            sample.append([("" if c is None else str(c)) for c in r[:len(header)]])
        # nb de lignes de données non vides
        n_data = sum(1 for r in data_rows if any(c not in (None, "") for c in r))
        tables_preview.append({
            "sheet": name, "position": pos, "header_row": hr,
            "columns": [h for h in header if h], "n_columns": len([h for h in header if h]),
            "n_rows": n_data, "sample": sample,
        })
    wb.close()

    preview = {
        "kind": kind,
        "n_tables": len(tables_preview),
        "tables": tables_preview,
        "is_reupload": DataSource.objects.filter(
            dataset_key=source.dataset_key, status=STATUS_COMMITTED
        ).exists(),
    }
    source.kind = kind
    source.preview = preview
    source.save(update_fields=["kind", "preview"])
    return preview


@transaction.atomic
def commit(source: DataSource, *, by: str = "") -> dict:
    """
    Enregistrement MANUEL : écrit les tables génériques + gère l'historique + alimente
    le typé (référentiel). Réupload d'un même dataset → nouvelle révision courante,
    anciennes conservées.
    """
    previous = (DataSource.objects
                .filter(dataset_key=source.dataset_key, status=STATUS_COMMITTED, is_current=True)
                .exclude(pk=source.pk)
                .order_by("-revision")
                .first())
    if previous:
        source.revision = previous.revision + 1
        source.supersedes = previous
        # L'ancienne reste en base (historique) mais n'est plus courante.
        DataSource.objects.filter(dataset_key=source.dataset_key, is_current=True).update(is_current=False)

    # (Ré)écriture des tables génériques de CETTE source.
    source.tables.all().delete()
    wb = openpyxl.load_workbook(source.file.path, data_only=True, read_only=True)
    total_records = 0
    for pos, name in enumerate(wb.sheetnames):
        ws = wb[name]
        rows = _sheet_matrix(ws)
        hr = _find_header_row(rows)
        header = [str(c).strip() if c is not None else "" for c in (rows[hr] if hr < len(rows) else ())]
        # colonnes nommées (ignore les colonnes sans en-tête)
        col_idx = [(i, h) for i, h in enumerate(header) if h]
        data_rows = rows[hr + 1:]

        table = DataTable.objects.create(
            source=source, name=name[:200], position=pos,
            n_cols=len(col_idx),
            n_rows=sum(1 for r in data_rows if any(c not in (None, "") for c in r)),
        )
        # Type inféré par colonne (échantillon).
        for cpos, (i, h) in enumerate(col_idx):
            sample_vals = [r[i] if i < len(r) else None for r in data_rows[:40]]
            DataColumn.objects.create(table=table, name=h[:255], position=cpos, dtype=_infer_dtype(sample_vals))
        # Lignes.
        records = []
        ri = 0
        for r in data_rows:
            if not any(c not in (None, "") for c in r):
                continue
            values = {h: (None if (i >= len(r) or r[i] is None) else str(r[i])) for i, h in col_idx}
            records.append(DataRecord(table=table, row_index=ri, values=values))
            ri += 1
        DataRecord.objects.bulk_create(records, batch_size=500)
        total_records += len(records)
    wb.close()

    # Hybride : si référentiel, alimente aussi les tables typées que le moteur lit.
    typed = None
    if source.kind == KIND_REFERENTIEL:
        report = ingest_workbook(
            source.file.path,
            label=f"{source.original_name} (r{source.revision})",
            source_filename=source.original_name,
        )
        typed = {"version_id": report.version_id, "total_ranges": report.total_ranges,
                 "config_loaded": report.config_loaded}

    source.status = STATUS_COMMITTED
    source.is_current = True
    source.committed_at = timezone.now()
    source.committed_by = by
    source.save(update_fields=["status", "is_current", "revision", "supersedes",
                               "committed_at", "committed_by"])

    return {
        "source_id": source.pk, "revision": source.revision,
        "superseded": previous.pk if previous else None,
        "tables": source.tables.count(), "records": total_records,
        "typed": typed,
    }


def _resync_referentiel(table: DataTable) -> dict | None:
    """
    Si la table appartient à un référentiel COURANT, re-dérive les plages typées que le
    moteur relit à partir des lignes ACTUELLES en base (après édition/suppression) →
    effet immédiat sur l'analyse. `None` sinon.
    """
    src = table.source
    if src.kind != KIND_REFERENTIEL or not src.is_current:
        return None
    version = (ReferentielVersion.objects
               .filter(is_active=True, source_filename=src.original_name)
               .order_by("-imported_at").first())
    if not version:
        return None
    header_names = [c.name for c in table.columns.all()]
    row_dicts = [r.values for r in table.records.all()]   # relit les valeurs à jour
    chain = BY_SHEET.get(table.name)
    if chain:
        return {"chaine": chain.code,
                "plages": rebuild_chain_from_records(version, chain, header_names, row_dicts)}
    if table.name == CALIBRATION_SHEET:
        return {"config": rebuild_config_from_records(version, header_names, row_dicts)}
    return None


@transaction.atomic
def update_records(table: DataTable, updates: list[dict] | None = None,
                   deletions: list[int] | None = None, *, by: str = "") -> dict:
    """
    Édition manuelle des lignes d'une table générique : CORRECTIONS (`updates` =
    [{id, values}]) et SUPPRESSIONS (`deletions` = [id]). On ne touche qu'aux colonnes
    connues d'une ligne (robustesse). Puis, si c'est un référentiel courant, on re-dérive
    les plages typées que le moteur relit. Tout est écrit en base (transaction atomique).
    """
    updates = updates or []
    deletions = [d for d in (deletions or []) if d]

    ids = [u.get("id") for u in updates if isinstance(u, dict) and u.get("id")]
    by_id = {r.id: r for r in table.records.filter(id__in=ids)}
    changed = 0
    for u in updates:
        rec = by_id.get(u.get("id")) if isinstance(u, dict) else None
        if not rec:
            continue
        merged = dict(rec.values)
        for k, v in (u.get("values") or {}).items():
            if k in merged:
                merged[k] = None if v in (None, "") else str(v)
        if merged != rec.values:
            rec.values = merged
            rec.save(update_fields=["values"])
            changed += 1

    deleted = 0
    if deletions:
        deleted = table.records.filter(id__in=deletions).delete()[0]

    typed = _resync_referentiel(table) if (changed or deleted) else None
    return {"changed": changed, "deleted": deleted, "typed": typed}


@transaction.atomic
def delete_source(source: DataSource) -> dict:
    """
    Supprime une source ET ses données (tables/colonnes/lignes en cascade). Pour un
    référentiel, retire aussi la version typée créée par ce commit (plages relues par le
    moteur) et réactive la version restante la plus récente. Si la source était courante,
    on promeut la révision précédente du même dataset.
    """
    kind = source.kind
    label = f"{source.original_name} (r{source.revision})"
    was_current = source.is_current
    key = source.dataset_key
    source.delete()  # cascade : DataTable / DataColumn / DataRecord

    typed_removed = 0
    if kind == KIND_REFERENTIEL:
        vqs = ReferentielVersion.objects.filter(label=label)
        was_active = vqs.filter(is_active=True).exists()
        typed_removed = vqs.count()
        vqs.delete()  # cascade : ReferenceRange
        if was_active:
            newest = ReferentielVersion.objects.order_by("-imported_at").first()
            if newest:
                ReferentielVersion.objects.exclude(pk=newest.pk).update(is_active=False)
                ReferentielVersion.objects.filter(pk=newest.pk).update(is_active=True)

    promoted = None
    if was_current:
        prev = (DataSource.objects
                .filter(dataset_key=key, status=STATUS_COMMITTED)
                .order_by("-revision").first())
        if prev:
            DataSource.objects.filter(pk=prev.pk).update(is_current=True)
            promoted = prev.pk

    return {"deleted": True, "typed_removed": typed_removed, "promoted": promoted}
