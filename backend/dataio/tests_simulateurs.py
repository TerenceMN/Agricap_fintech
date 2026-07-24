"""Tests des simulateurs manquants construits par `build_simulateurs_manquants`.

Dette CLAUDE.md §6 : « créer les simulateurs manquants (CAFE_ARABICA, MANIOC,
RIZ) ». Ce qui est vérifié ici, c'est que les trois classeurs empruntent
EXACTEMENT le chemin des quatorze existants — aucune migration, aucun schéma,
aucun `if` de filière dans le moteur :

  classeur → `dataio.services.inspect` (détection SIMULATEUR)
           → `dataio.services.commit`  (tables/colonnes/lignes génériques)
           → `credits.referentiel_loader.charger_depuis_simulateur`

et que ce qui n'est porté par aucun référentiel (rendement, prix, plages de
vraisemblance) reste VIDE plutôt que deviné.
"""
from __future__ import annotations

import os
import shutil
import tempfile
from decimal import Decimal

import openpyxl
from django.conf import settings
from django.core.files import File
from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from unittest import skipUnless

from credits.referentiel_loader import ReferentielIntrouvable, charger_depuis_simulateur
from dataio import services
from dataio.management.commands.build_simulateurs_manquants import (
    DEFAULT_SUBDIR, FILIERES_MANQUANTES, SUPERFICIE_REFERENCE,
    _nom_fichier, repartir_couts,
)
from dataio.models import KIND_SIMULATEUR, DataSource
from reference_data.models import MODULE_WEIGHT_KEYS, ReferenceFileUpload, ValueChain

MODELES_DIR = os.path.join(settings.DOCUMENT_EXCEL_DIR, DEFAULT_SUBDIR)

#: Le référentiel institutionnel des trois filières concernées, tel qu'il est
#: chargé en base (`reference_data/fixtures/value_chains_initial.py`). Les
#: classeurs n'ont pas d'autre source de chiffres.
CHAINES = {
    "RIZ": {
        "label": "Riz", "cycle_months": 6,
        "cost_per_hectare_usd": "1500.00", "cost_per_hectare_cdf": "4200000.00",
        "module_weights": {
            "semences": 20, "mecanisation": 18, "maindoeuvre": 28, "equipements": 8,
            "postrecolte": 12, "logistique": 8, "commercialisation": 3, "reserve": 3,
        },
        "risk_factor": "1.050", "min_score_required": 58, "base_rate": "7.00",
        "harvest_months": [2, 8], "eligible_guarantees": ["epargne", "morale"],
    },
    "MANIOC": {
        "label": "Manioc", "cycle_months": 12,
        "cost_per_hectare_usd": "900.00", "cost_per_hectare_cdf": "2520000.00",
        "module_weights": {
            "semences": 20, "mecanisation": 10, "maindoeuvre": 35, "equipements": 5,
            "postrecolte": 10, "logistique": 10, "commercialisation": 5, "reserve": 5,
        },
        "risk_factor": "0.850", "min_score_required": 50, "base_rate": "8.00",
        "harvest_months": [11, 12], "eligible_guarantees": ["epargne", "morale"],
    },
    "CAFE_ARABICA": {
        "label": "Café Arabica", "cycle_months": 9,
        "cost_per_hectare_usd": "4800.00", "cost_per_hectare_cdf": "13440000.00",
        "module_weights": {
            "semences": 15, "mecanisation": 10, "maindoeuvre": 30, "equipements": 12,
            "postrecolte": 15, "logistique": 8, "commercialisation": 5, "reserve": 5,
        },
        "risk_factor": "0.900", "min_score_required": 60, "base_rate": "6.00",
        "harvest_months": [3, 4], "eligible_guarantees": ["epargne", "morale"],
    },
}

modeles_presents = skipUnless(
    os.path.isdir(MODELES_DIR),
    f"Classeurs modèles absents ({MODELES_DIR}) — la forme d'un simulateur n'est "
    f"jamais retapée dans le code, elle vient du classeur institutionnel.",
)


def _seed_referentiel(codes=CHAINES.keys()) -> dict[str, ValueChain]:
    upload = ReferenceFileUpload.objects.create(
        file="reference_data/fixture.xlsx",
        file_type=ReferenceFileUpload.FileType.VALUE_CHAINS,
        version="test-v1", uploaded_by="maker-1", activated_by="checker-2",
        activated_at=timezone.now(), status=ReferenceFileUpload.Status.ACTIVE,
    )
    return {
        code: ValueChain.objects.create(source_file=upload, code=code, **CHAINES[code])
        for code in codes
    }


class RepartitionCoutsTests(TestCase):
    """La ventilation des coûts est une propriété, pas une estimation."""

    def test_somme_des_modules_egale_exactement_le_cout_par_hectare(self):
        chains = _seed_referentiel()
        for code, chain in chains.items():
            with self.subTest(code=code):
                montants = repartir_couts(chain)
                total = sum(montants.values(), Decimal("0.00"))
                # Σ feuille 5 = coût/ha au centime près (principe 5) : un total
                # « à peu près » se paierait en contrôle de cohérence rouge sur
                # chaque dossier de la filière.
                self.assertEqual(
                    total, Decimal(chain.cost_per_hectare_usd) * SUPERFICIE_REFERENCE,
                    f"{code} : Σ modules ≠ coût/ha",
                )

    def test_chaque_module_recoit_son_poids_du_referentiel(self):
        chain = _seed_referentiel(["RIZ"])["RIZ"]
        montants = repartir_couts(chain)
        cout_ha = Decimal(chain.cost_per_hectare_usd)
        for module in MODULE_WEIGHT_KEYS:
            attendu = cout_ha * Decimal(str(chain.module_weights[module])) / Decimal(100)
            self.assertAlmostEqual(
                montants[module], attendu, delta=Decimal("0.01"),
                msg=f"module {module} : ventilation hors grille du référentiel",
            )

    def test_aucun_float_dans_la_repartition(self):
        chain = _seed_referentiel(["CAFE_ARABICA"])["CAFE_ARABICA"]
        for montant in repartir_couts(chain).values():
            self.assertIsInstance(montant, Decimal)


@modeles_presents
class ConstructionClasseursTests(TestCase):
    """Les trois classeurs manquants, écrits et relus comme les quatorze autres."""

    def setUp(self) -> None:
        self.chains = _seed_referentiel()
        self.outdir = tempfile.mkdtemp(prefix="agricap-sim-")
        # Les classeurs modèles doivent être visibles depuis le dossier de sortie :
        # la commande y cherche la FORME (feuilles, en-têtes, formules).
        for spec in FILIERES_MANQUANTES.values():
            src = os.path.join(MODELES_DIR, spec["modele"])
            if os.path.isfile(src):
                shutil.copy2(src, os.path.join(self.outdir, spec["modele"]))
        call_command("build_simulateurs_manquants", output=self.outdir, verbosity=0)

    def tearDown(self) -> None:
        shutil.rmtree(self.outdir, ignore_errors=True)

    def _chemin(self, code: str) -> str:
        return os.path.join(self.outdir, _nom_fichier(FILIERES_MANQUANTES[code]))

    def _ingerer(self, code: str) -> DataSource:
        """Exactement le chemin de `ingest_simulateurs` : inspect puis commit."""
        path = self._chemin(code)
        name = os.path.basename(path)
        src = DataSource(original_name=name, dataset_key=services.dataset_key(name),
                         uploaded_by="test")
        with open(path, "rb") as fh:
            src.file.save(name, File(fh), save=False)
        src.save()
        services.inspect(src)
        services.commit(src, by="test")
        return src

    # ── Le classeur existe et est reconnu ────────────────────────────────────

    def test_les_trois_classeurs_sont_ecrits(self):
        for code in FILIERES_MANQUANTES:
            with self.subTest(code=code):
                self.assertTrue(os.path.isfile(self._chemin(code)))

    def test_chaque_classeur_est_detecte_comme_simulateur(self):
        for code in FILIERES_MANQUANTES:
            with self.subTest(code=code):
                wb = openpyxl.load_workbook(self._chemin(code), read_only=True)
                sheets = wb.sheetnames
                wb.close()
                self.assertGreaterEqual(len(sheets), 15)
                self.assertEqual(services.detect_kind(sheets), KIND_SIMULATEUR)

    def test_le_nom_suit_la_convention_lue_par_le_loader(self):
        """Le nom porte le code de CHAÎNE (01–14), pas une quinzième nomenclature."""
        for code, spec in FILIERES_MANQUANTES.items():
            with self.subTest(code=code):
                self.assertIn(spec["numero"], {c.code for c in _chaines_referentiel_v3()})

    # ── Ingestion générique : aucune migration, aucun schéma ─────────────────

    def test_ingestion_ecrit_les_tables_generiques(self):
        src = self._ingerer("RIZ")
        self.assertEqual(src.kind, KIND_SIMULATEUR)
        self.assertTrue(src.is_current)
        noms = set(src.tables.values_list("name", flat=True))
        for feuille in ("2_Identification_Projet", "3_Parametres_Techniques",
                        "5_Synthese_Besoins"):
            self.assertIn(feuille, noms)

    # ── Lecture du référentiel : c'est là qu'est la valeur de la dette ───────

    def test_le_referentiel_se_charge_pour_les_trois_filieres(self):
        for code in FILIERES_MANQUANTES:
            with self.subTest(code=code):
                src = self._ingerer(code)
                try:
                    spec = charger_depuis_simulateur(src)
                except ReferentielIntrouvable as exc:      # pragma: no cover
                    self.fail(f"{code} : référentiel illisible — {exc}")
                self.assertEqual(spec["unite_reference"], "ha")
                self.assertEqual(spec["devise"], "USD")
                self.assertEqual(spec["source"], "indicatif")

    def test_les_couts_par_module_sont_ceux_du_referentiel_valuechain(self):
        """La fiabilité technique (25 % de la note) devient calculable, sur les
        chiffres de l'institution et non sur une répartition estimée."""
        for code, chain in self.chains.items():
            with self.subTest(code=code):
                spec = charger_depuis_simulateur(self._ingerer(code))
                couts = spec["couts_modules"]
                self.assertEqual(set(couts), set(MODULE_WEIGHT_KEYS))
                attendus = repartir_couts(chain)
                for module in MODULE_WEIGHT_KEYS:
                    self.assertEqual(
                        Decimal(couts[module]["ref"]),
                        attendus[module] / SUPERFICIE_REFERENCE,
                        f"{code}/{module} : coût de référence hors grille",
                    )

    def test_le_total_lu_egale_le_cout_par_hectare_du_referentiel(self):
        for code, chain in self.chains.items():
            with self.subTest(code=code):
                spec = charger_depuis_simulateur(self._ingerer(code))
                lignage = spec["_lignage"]
                self.assertEqual(Decimal(lignage["quantiteReference"]),
                                 Decimal(SUPERFICIE_REFERENCE))
                self.assertEqual(Decimal(lignage["totalCycleLu"]),
                                 Decimal(chain.cost_per_hectare_usd))

    # ── Ce qui n'est porté par aucun référentiel reste vide ──────────────────

    def test_le_rendement_reste_vide_faute_de_reference(self):
        """Aucune table AGRICAP ne porte de rendement pour ces filières : le
        classeur le laisse vide plutôt que de fabriquer un DSCR."""
        for code in FILIERES_MANQUANTES:
            with self.subTest(code=code):
                spec = charger_depuis_simulateur(self._ingerer(code))
                self.assertEqual(spec["rendement_ref"], {})

    def test_aucune_plage_de_vraisemblance_inventee(self):
        wb = openpyxl.load_workbook(self._chemin("MANIOC"), data_only=True, read_only=True)
        ws = wb["18_Controles_Vraisemblance"]
        bornes = [(r[2], r[3]) for r in ws.iter_rows(min_row=5, max_row=10, values_only=True)]
        wb.close()
        self.assertTrue(all(mini is None and maxi is None for mini, maxi in bornes),
                        f"plages écrites sans référence : {bornes}")

    def test_la_filiere_porte_le_libelle_du_referentiel(self):
        spec = charger_depuis_simulateur(self._ingerer("CAFE_ARABICA"))
        self.assertIn("Café Arabica", spec["filiere"])

    # ── Garde-fous de la commande ────────────────────────────────────────────

    def test_sans_referentiel_actif_aucun_classeur_n_est_fabrique(self):
        """Pas de coût deviné : sans `ValueChain` actif, la commande s'abstient."""
        ValueChain.objects.all().delete()
        vide = tempfile.mkdtemp(prefix="agricap-sim-vide-")
        try:
            for spec in FILIERES_MANQUANTES.values():
                src = os.path.join(MODELES_DIR, spec["modele"])
                if os.path.isfile(src):
                    shutil.copy2(src, os.path.join(vide, spec["modele"]))
            call_command("build_simulateurs_manquants", output=vide, verbosity=0)
            for spec in FILIERES_MANQUANTES.values():
                self.assertFalse(os.path.exists(os.path.join(vide, _nom_fichier(spec))))
        finally:
            shutil.rmtree(vide, ignore_errors=True)

    def test_un_classeur_existant_n_est_pas_ecrase_sans_force(self):
        cible = self._chemin("RIZ")
        avant = os.path.getmtime(cible)
        call_command("build_simulateurs_manquants", output=self.outdir, verbosity=0)
        self.assertEqual(os.path.getmtime(cible), avant)


def _chaines_referentiel_v3():
    from referentiel.chains import CHAINS
    return CHAINS
