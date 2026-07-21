"""Tests du mécanisme de templates versionnés (principe 11) — `dataio.FileTemplate`.

Couverture :
  - dérivation automatique du schéma à l'upload/activation (feuilles, colonnes, rubriques) ;
  - maker-checker : activation par un checker ≠ maker ; refus si maker == checker ;
  - le précédent template actif passe `archived` ;
  - validation structurelle d'un fichier client CONTRE le template actif (et non un schéma
    codé en dur), avec enregistrement de `templateId + version` ;
  - refus explicite `TEMPLATE_NOT_CONFIGURED` sans template actif ;
  - permissions : capacité `config` requise pour upload/activation/liste.
"""
from __future__ import annotations

import io
import os
import tempfile

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile

from common.testing import AuthedAPITestCase
from dataio import services_templates as tpl_svc
from dataio.models import FileTemplate, KIND_FEUILLE_BESOINS

HEADER_4 = [
    "N°", "Rubrique", "Description détaillée", "Unité", "Quantité", "Coût unitaire",
    "Fréquence", "Montant total", "Période du cycle", "Financement souhaité",
    "Observations", "Validation analyste (réservé)",
]
HEADER_5 = ["Rubrique", "Total rubrique", "Part du total", "Commentaire"]
RUBRIQUES = [
    "Semences & Intrants", "Opérations mécanisées", "Main d'œuvre",
    "Équipement & petit matériel", "Récolte & post-récolte", "Logistique",
    "Commercialisation", "Réserve d'exploitation",
]


def build_fb(
    *,
    header4: list | None = None,
    header5: list | None = None,
    omit_sheets: tuple[str, ...] = (),
    omit_rubriques: tuple[str, ...] = (),
) -> bytes:
    """Construit un classeur de feuille de besoins (feuilles 4 & 5) et renvoie ses octets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if "4_Besoins_Financiers" not in omit_sheets:
        ws4 = wb.create_sheet("4_Besoins_Financiers")
        ws4.append(["4. Besoins financiers détaillés du cycle"])
        ws4.append(["Saisir uniquement les cellules jaunes."])
        ws4.append([])
        ws4.append(header4 if header4 is not None else HEADER_4)
        ws4.append([1, "Semences & Intrants", "Semences", "kg", 50, 12, 1, 600,
                    "Préparation", "Crédit", "", None])

    if "5_Synthese_Besoins" not in omit_sheets:
        ws5 = wb.create_sheet("5_Synthese_Besoins")
        ws5.append(["5. Synthèse des besoins financiers"])
        ws5.append(["Feuille calculée automatiquement — ne rien saisir."])
        ws5.append(header5 if header5 is not None else HEADER_5)
        for label in RUBRIQUES:
            if label in omit_rubriques:
                continue
            ws5.append([label, 0, 0, None])
        ws5.append(["TOTAL GÉNÉRAL", 600, 1, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_file(data: bytes, name: str = "template.xlsx") -> SimpleUploadedFile:
    return SimpleUploadedFile(
        name, data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _write_tmp(data: bytes) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(data)
    tmp.close()
    return tmp.name


def _seed_active(*, maker="maker-1", checker="checker-2") -> FileTemplate:
    tpl = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by=maker)
    tpl_svc.activate_template(tpl, activator_sub=checker)
    return tpl


def _codes(errors) -> set[str]:
    return {e["code"] for e in errors}


class SchemaDerivationTests(AuthedAPITestCase):
    def test_upload_derives_schema_sheets_columns_rubriques(self):
        tpl = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="maker-1")
        self.assertEqual(tpl.status, FileTemplate.Status.PENDING)
        self.assertEqual(tpl.version, 1)
        self.assertTrue(tpl.sha256)
        schema = tpl.schema
        self.assertIn("4_Besoins_Financiers", schema["sheet_names"])
        self.assertIn("5_Synthese_Besoins", schema["sheet_names"])
        self.assertEqual(schema["synthesis_sheet"], "5_Synthese_Besoins")
        # Rubriques dérivées de la feuille de synthèse (TOTAL GÉNÉRAL exclu).
        self.assertEqual(len(schema["rubriques"]), 8)
        self.assertIn("Semences & Intrants", schema["rubriques"])
        self.assertNotIn("TOTAL GÉNÉRAL", schema["rubriques"])
        # Colonnes de la feuille 4 dérivées dans l'ordre.
        s4 = next(s for s in schema["sheets"] if s["name"] == "4_Besoins_Financiers")
        self.assertEqual(s4["columns"][:3], ["N°", "Rubrique", "Description détaillée"])

    def test_second_upload_bumps_version(self):
        tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="maker-1")
        tpl2 = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="maker-1")
        self.assertEqual(tpl2.version, 2)

    def test_upload_rejects_non_xlsx(self):
        with self.assertRaises(tpl_svc.TemplateUploadError) as ctx:
            tpl_svc.upload_template(_upload_file(b"pas un xlsx", name="t.xlsm"))
        self.assertEqual(ctx.exception.code, "EXTENSION_INVALIDE")


class MakerCheckerTests(AuthedAPITestCase):
    def test_activation_by_checker_archives_previous(self):
        v1 = _seed_active(maker="m1", checker="c1")
        self.assertEqual(v1.status, FileTemplate.Status.ACTIVE)

        v2 = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="m1")
        tpl_svc.activate_template(v2, activator_sub="c1")

        v1.refresh_from_db()
        v2.refresh_from_db()
        self.assertEqual(v2.status, FileTemplate.Status.ACTIVE)
        self.assertEqual(v1.status, FileTemplate.Status.ARCHIVED)
        self.assertEqual(v2.supersedes_id, v1.pk)
        self.assertEqual(tpl_svc.active_template(KIND_FEUILLE_BESOINS).pk, v2.pk)

    def test_activation_refused_when_maker_equals_checker(self):
        tpl = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="same-sub")
        with self.assertRaises(tpl_svc.TemplateActivationError) as ctx:
            tpl_svc.activate_template(tpl, activator_sub="same-sub")
        self.assertEqual(ctx.exception.code, "MAKER_EGAL_CHECKER")
        tpl.refresh_from_db()
        self.assertEqual(tpl.status, FileTemplate.Status.PENDING)

    def test_activation_refused_when_not_pending(self):
        tpl = _seed_active(maker="m1", checker="c1")
        with self.assertRaises(tpl_svc.TemplateActivationError) as ctx:
            tpl_svc.activate_template(tpl, activator_sub="c9")
        self.assertEqual(ctx.exception.code, "STATUT_INVALIDE")

    def test_activation_writes_audit(self):
        from audit.models import AuditEntry
        _seed_active(maker="m1", checker="c1")
        self.assertTrue(AuditEntry.objects.filter(
            action="dataio.template.activate", entity_type="FileTemplate").exists())


class ValidateStructureTests(AuthedAPITestCase):
    def test_no_active_template_returns_template_not_configured(self):
        path = _write_tmp(build_fb())
        try:
            errors, ref = tpl_svc.validate_structure(path)
        finally:
            os.unlink(path)
        self.assertEqual(_codes(errors), {"TEMPLATE_NOT_CONFIGURED"})
        self.assertEqual(ref, {})

    def test_conform_file_passes_and_records_template_ref(self):
        tpl = _seed_active()
        path = _write_tmp(build_fb())
        try:
            errors, ref = tpl_svc.validate_structure(path)
        finally:
            os.unlink(path)
        self.assertEqual(errors, [])
        self.assertEqual(ref, {"templateId": tpl.pk, "version": tpl.version})

    def test_missing_sheet_detected_against_active_template(self):
        _seed_active()
        path = _write_tmp(build_fb(omit_sheets=("5_Synthese_Besoins",)))
        try:
            errors, ref = tpl_svc.validate_structure(path)
        finally:
            os.unlink(path)
        self.assertIn("FEUILLE_MANQUANTE", _codes(errors))
        self.assertTrue(ref)  # templateId + version toujours renvoyés

    def test_missing_column_detected(self):
        _seed_active()
        header = list(HEADER_4)
        header[4] = "Nbre"  # « Quantité » renommé → plus reconnu
        path = _write_tmp(build_fb(header4=header))
        try:
            errors, _ref = tpl_svc.validate_structure(path)
        finally:
            os.unlink(path)
        self.assertIn("COLONNE_MANQUANTE", _codes(errors))

    def test_missing_rubrique_detected(self):
        _seed_active()
        path = _write_tmp(build_fb(omit_rubriques=("Logistique",)))
        try:
            errors, _ref = tpl_svc.validate_structure(path)
        finally:
            os.unlink(path)
        self.assertIn("RUBRIQUE_MANQUANTE", _codes(errors))

    def test_required_sheets_scoping(self):
        _seed_active()
        # On n'exige QUE la feuille 4 : l'absence de la feuille 5 ne doit rien lever.
        path = _write_tmp(build_fb(omit_sheets=("5_Synthese_Besoins",)))
        try:
            errors, _ref = tpl_svc.validate_structure(
                path, required_sheets=["4_Besoins_Financiers"])
        finally:
            os.unlink(path)
        self.assertEqual(errors, [])

    def test_serve_active_returns_bytes(self):
        tpl = _seed_active()
        data, name, ref = tpl_svc.serve_active()
        self.assertGreater(len(data), 0)
        self.assertEqual(ref, {"templateId": tpl.pk, "version": tpl.version})

    def test_serve_active_raises_without_template(self):
        with self.assertRaises(tpl_svc.TemplateNotConfigured):
            tpl_svc.serve_active()


class TemplateApiTests(AuthedAPITestCase):
    def test_upload_activate_list_flow(self):
        # Maker upload (capacité config).
        self.login(role="admin_it", sub="maker-a")
        res = self.client.post(
            "/api/dataio/templates/upload",
            {"file": _upload_file(build_fb())}, format="multipart")
        self.assertEqual(res.status_code, 201)
        tid = res.data["id"]
        self.assertEqual(res.data["status"], "pending")
        self.assertEqual(len(res.data["rubriques"]), 8)

        # Maker ne peut pas activer son propre template.
        refused = self.client.post(f"/api/dataio/templates/{tid}/activate")
        self.assertEqual(refused.status_code, 409)
        self.assertEqual(refused.data["errors"][0]["code"], "MAKER_EGAL_CHECKER")

        # Checker ≠ maker active.
        self.login(role="admin_it", sub="checker-b")
        ok = self.client.post(f"/api/dataio/templates/{tid}/activate")
        self.assertEqual(ok.status_code, 200)
        self.assertEqual(ok.data["status"], "active")

        # Liste : template actif exposé.
        listed = self.client.get("/api/dataio/templates/")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.data["active"]["id"], tid)

    def test_upload_requires_config_capability(self):
        self.login(role="agent_terrain", sub="no-config")  # read/create, pas config
        res = self.client.post(
            "/api/dataio/templates/upload",
            {"file": _upload_file(build_fb())}, format="multipart")
        self.assertEqual(res.status_code, 403)

    def test_list_requires_config_capability(self):
        self.login(role="agent_terrain", sub="no-config")
        res = self.client.get("/api/dataio/templates/")
        self.assertEqual(res.status_code, 403)


class TemplateDetailTests(AuthedAPITestCase):
    """Le CHECKER (≠ maker) doit voir le schéma dérivé complet et le diff AU
    RECHARGEMENT de l'écran, pas seulement dans la réponse d'upload du maker —
    sans quoi il activerait à l'aveugle (CLAUDE.md §7.1.5)."""

    def test_detail_exposes_full_schema_and_diff_to_checker(self):
        v1 = _seed_active(maker="m1", checker="c1")
        # v2 ajoute une rubrique et retire une feuille par rapport à l'actif.
        v2 = tpl_svc.upload_template(
            _upload_file(build_fb(omit_rubriques=("Logistique",))), uploaded_by="m1")

        # Le checker, qui n'a pas fait l'upload, ouvre l'écran.
        self.login(role="admin_it", sub="c1")
        res = self.client.get(f"/api/dataio/templates/{v2.pk}")
        self.assertEqual(res.status_code, 200)

        # Schéma dérivé COMPLET (colonnes/types/row_labels), pas seulement le résumé.
        self.assertIn("schema", res.data)
        s4 = next(s for s in res.data["schema"]["sheets"]
                  if s["name"] == "4_Besoins_Financiers")
        self.assertIn("columns", s4)
        self.assertIn("types", s4)

        # Diff calculé SERVEUR vs le template actif (la question du checker).
        self.assertIn("Logistique", res.data["diff"]["rubriquesRemoved"])
        self.assertEqual(res.data["diffBaseline"]["id"], v1.pk)
        self.assertEqual(res.data["diffBaseline"]["relation"], "active")

    def test_detail_of_active_diffs_against_superseded(self):
        v1 = _seed_active(maker="m1", checker="c1")
        v2 = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="m1")
        tpl_svc.activate_template(v2, activator_sub="c1")

        self.login(role="admin_it", sub="c1")
        res = self.client.get(f"/api/dataio/templates/{v2.pk}")
        self.assertEqual(res.status_code, 200)
        # Trace historique : ce qui a changé au moment de SON activation.
        self.assertEqual(res.data["diffBaseline"]["id"], v1.pk)
        self.assertEqual(res.data["diffBaseline"]["relation"], "supersedes")

    def test_detail_without_baseline_reports_no_previous(self):
        tpl = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="m1")
        self.login(role="admin_it", sub="c1")
        res = self.client.get(f"/api/dataio/templates/{tpl.pk}")
        self.assertEqual(res.status_code, 200)
        self.assertIsNone(res.data["diffBaseline"]["id"])
        self.assertFalse(res.data["diff"]["hasPrevious"])

    def test_detail_404_when_unknown(self):
        self.login(role="admin_it", sub="c1")
        self.assertEqual(self.client.get("/api/dataio/templates/9999").status_code, 404)

    def test_detail_requires_config_capability(self):
        tpl = tpl_svc.upload_template(_upload_file(build_fb()), uploaded_by="m1")
        self.login(role="agent_terrain", sub="no-config")
        res = self.client.get(f"/api/dataio/templates/{tpl.pk}")
        self.assertEqual(res.status_code, 403)

    def test_detail_route_does_not_shadow_upload(self):
        """`templates/<int:pk>` ne doit pas capturer `templates/upload`."""
        self.login(role="admin_it", sub="maker-a")
        res = self.client.post(
            "/api/dataio/templates/upload",
            {"file": _upload_file(build_fb())}, format="multipart")
        self.assertEqual(res.status_code, 201)
