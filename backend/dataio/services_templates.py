"""
Principe 11 — templates de fichiers versionnés (maker-checker) et validation
structurelle des fichiers client CONTRE le template actif, jamais contre un schéma
codé en dur.

Ce module est le point d'exposition côté `dataio` que les autres apps (en pratique
`credits`, qui possède la feuille de besoins) branchent :

    - `validate_structure(path, kind)` valide la FORME d'un fichier client contre le
      schéma dérivé du template actif, et renvoie `templateId + version` à journaliser ;
    - `serve_active(kind)` renvoie les octets du template actif à télécharger (le fichier
      téléchargé = le fichier contre lequel on valide) ;
    - sans template actif : `TEMPLATE_NOT_CONFIGURED` (jamais de validation « best effort »).

Le schéma n'est pas maintenu à la main : `derive_schema()` le calcule depuis le fichier
lui-même, avec EXACTEMENT la même lecture (`services.sheet_rows`) que l'ingestion — on ne
valide jamais une lecture différente de celle qui sera écrite en base.
"""
from __future__ import annotations

from django.core.files import File
from django.db import transaction
from django.utils import timezone

import openpyxl

from . import services as svc
from .models import FileTemplate, KIND_FEUILLE_BESOINS

#: Taille maximale d'un template (cohérent avec les uploads client, CLAUDE.md §5).
MAX_TEMPLATE_BYTES = 5 * 1024 * 1024


# ── Exceptions ────────────────────────────────────────────────────────────────

class TemplateNotConfigured(Exception):
    """Aucun template actif pour un `kind` : l'upload/la validation client est refusé."""

    code = "TEMPLATE_NOT_CONFIGURED"

    def __init__(self, kind: str) -> None:
        self.kind = kind
        self.message = (
            f"Aucun template actif n'est configuré pour « {kind} ». "
            f"Un administrateur doit téléverser puis activer un template de référence "
            f"avant que des fichiers puissent être déposés ou validés."
        )
        super().__init__(self.message)


class TemplateUploadError(Exception):
    """Upload refusé (extension, taille, classeur illisible)."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        self.message = message
        super().__init__(message)


class TemplateActivationError(Exception):
    """Activation refusée (statut, maker == checker)."""

    def __init__(self, code: str, message: str) -> None:
        self.code = code
        self.message = message
        super().__init__(message)


# ── Dérivation du schéma ──────────────────────────────────────────────────────

def derive_schema(file_path: str) -> dict:
    """Dérive le schéma attendu d'un classeur : feuilles, colonnes (ordre), types inférés,
    libellés de lignes, et la liste des rubriques (première colonne de la feuille de
    synthèse). Générique — aucune connaissance métier codée en dur.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheets: list[dict] = []
        synthesis_sheet: str | None = None

        for pos, name in enumerate(wb.sheetnames):
            header, data = svc.sheet_rows(wb[name])
            columns = [h for h in header if h]

            types: dict[str, str] = {}
            for idx, h in enumerate(header):
                if not h:
                    continue
                sample = [(r[idx] if idx < len(r) else None) for r in data[:40]]
                types[h] = svc._infer_dtype(sample)

            first_idx = next((i for i, h in enumerate(header) if h), None)
            row_labels: list[str] = []
            if first_idx is not None:
                for r in data:
                    v = r[first_idx] if first_idx < len(r) else None
                    s = "" if v is None else str(v).strip()
                    if s:
                        row_labels.append(s)

            sheets.append({
                "name": name, "position": pos,
                "columns": columns, "n_columns": len(columns),
                "types": types, "row_labels": row_labels,
            })
            if synthesis_sheet is None and "synth" in svc._norm(name):
                synthesis_sheet = name

        rubriques: list[str] = []
        if synthesis_sheet is not None:
            sdef = next(s for s in sheets if s["name"] == synthesis_sheet)
            for lbl in sdef["row_labels"]:
                if svc._norm(lbl).startswith("total"):
                    continue
                if lbl not in rubriques:
                    rubriques.append(lbl)

        return {
            "sheets": sheets,
            "sheet_names": list(wb.sheetnames),
            "synthesis_sheet": synthesis_sheet,
            "rubriques": rubriques,
            "derived_at": timezone.now().isoformat(),
        }
    finally:
        wb.close()


def diff_schema(old: dict | None, new: dict | None) -> dict:
    """Diff lisible entre deux schémas (pour l'aperçu maker avant activation)."""
    old_sheets = {s["name"]: s for s in (old or {}).get("sheets", [])}
    new_sheets = {s["name"]: s for s in (new or {}).get("sheets", [])}
    columns_changed = [
        n for n in new_sheets
        if n in old_sheets and old_sheets[n].get("columns") != new_sheets[n].get("columns")
    ]
    old_rub = (old or {}).get("rubriques", [])
    new_rub = (new or {}).get("rubriques", [])
    return {
        "sheetsAdded": [n for n in new_sheets if n not in old_sheets],
        "sheetsRemoved": [n for n in old_sheets if n not in new_sheets],
        "sheetsColumnsChanged": columns_changed,
        "rubriquesAdded": [r for r in new_rub if r not in old_rub],
        "rubriquesRemoved": [r for r in old_rub if r not in new_rub],
        "hasPrevious": bool(old_sheets),
    }


# ── Lecture du template actif ─────────────────────────────────────────────────

def active_template(kind: str = KIND_FEUILLE_BESOINS) -> FileTemplate | None:
    """Le template ACTIF pour un `kind`, ou `None`."""
    return (FileTemplate.objects
            .filter(kind=kind, status=FileTemplate.Status.ACTIVE)
            .order_by("-activated_at", "-version").first())


def active_schema(kind: str = KIND_FEUILLE_BESOINS) -> dict:
    """Schéma du template actif. Lève `TemplateNotConfigured` s'il n'y en a pas."""
    tpl = active_template(kind)
    if tpl is None:
        raise TemplateNotConfigured(kind)
    return tpl.schema or {}


def template_ref(tpl: FileTemplate) -> dict:
    """Référence minimale à journaliser dans un rapport de validation (principe 11)."""
    return {"templateId": tpl.pk, "version": tpl.version}


# ── Validation structurelle d'un fichier client contre le template actif ──────

def validate_structure(
    path: str,
    kind: str = KIND_FEUILLE_BESOINS,
    required_sheets: list[str] | None = None,
) -> tuple[list[dict], dict]:
    """Valide la FORME d'un fichier client contre le schéma du template ACTIF.

    Renvoie `(errors, template_ref)` :
      - `errors` : liste de `{code, message}` (→ 422), `[]` si la forme est conforme ;
        `TEMPLATE_NOT_CONFIGURED` si aucun template n'est actif ;
      - `template_ref` : `{templateId, version}` du template utilisé (à enregistrer dans
        le rapport de validation) — `{}` si aucun template actif.

    `required_sheets` restreint les feuilles exigées (les colonnes de ces feuilles sont
    vérifiées). `None` → toutes les feuilles du template sont exigées.

    Cette fonction NE fait que la validation STRUCTURELLE (N1 : feuilles, colonnes,
    rubriques présentes). La cohérence interne (Σ, totaux — N2) reste à la charge de
    l'appelant métier.
    """
    tpl = active_template(kind)
    if tpl is None:
        return [{"code": TemplateNotConfigured.code,
                 "message": TemplateNotConfigured(kind).message}], {}

    ref = template_ref(tpl)
    schema = tpl.schema or {}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        return [{"code": "CLASSEUR_ILLISIBLE",
                 "message": f"Le classeur n'a pas pu être ouvert : {exc}"}], ref

    try:
        client_sheets: dict[str, tuple[list[str], list]] = {}
        for name in wb.sheetnames:
            header, data = svc.sheet_rows(wb[name])
            client_sheets[svc._norm(name)] = ([h for h in header if h], data)

        want = {svc._norm(s) for s in required_sheets} if required_sheets else None
        errors: list[dict] = []

        for sdef in schema.get("sheets", []):
            sname = sdef["name"]
            nkey = svc._norm(sname)
            if want is not None and nkey not in want:
                continue
            if nkey not in client_sheets:
                errors.append({
                    "code": "FEUILLE_MANQUANTE",
                    "message": (f"La feuille « {sname} » est absente du classeur. "
                                f"Utilisez le modèle officiel AGRICAP sans renommer "
                                f"les feuilles (template v{tpl.version})."),
                })
                continue
            client_cols = [svc._norm(c) for c in client_sheets[nkey][0]]
            for col in sdef.get("columns", []):
                nc = svc._norm(col)
                if not nc:
                    continue
                matched = any(
                    nc == cc or (len(nc) >= 3 and (nc in cc or cc in nc))
                    for cc in client_cols
                )
                if not matched:
                    errors.append({
                        "code": "COLONNE_MANQUANTE",
                        "message": (f"Feuille « {sname} » : colonne « {col} » introuvable. "
                                    f"En-têtes lus : "
                                    f"{', '.join(client_sheets[nkey][0]) or '(aucun)'}."),
                    })

        # Rubriques (dérivées de la feuille de synthèse du template).
        syn = schema.get("synthesis_sheet")
        rubriques = schema.get("rubriques", [])
        syn_required = want is None or (syn is not None and svc._norm(syn) in want)
        if syn and rubriques and syn_required and svc._norm(syn) in client_sheets:
            header, data = client_sheets[svc._norm(syn)]
            first_idx = 0
            present = set()
            for r in data:
                v = r[first_idx] if first_idx < len(r) else None
                s = "" if v is None else str(v).strip()
                if s:
                    present.add(svc._norm(s))
            for rub in rubriques:
                if svc._norm(rub) not in present:
                    errors.append({
                        "code": "RUBRIQUE_MANQUANTE",
                        "message": (f"Feuille « {syn} » : la rubrique « {rub} » du template "
                                    f"actif est absente. Toutes les rubriques du modèle "
                                    f"doivent figurer, même à 0."),
                    })

        return errors, ref
    finally:
        wb.close()


# ── Service de téléchargement (le fichier servi = celui contre lequel on valide) ──

def serve_active(kind: str = KIND_FEUILLE_BESOINS) -> tuple[bytes, str, dict]:
    """Octets + nom + `{templateId, version}` du template actif.

    Lève `TemplateNotConfigured` s'il n'y a pas de template actif — l'appelant renvoie
    alors le code d'erreur explicite, jamais un fichier « best effort ».
    """
    tpl = active_template(kind)
    if tpl is None:
        raise TemplateNotConfigured(kind)
    with tpl.file.open("rb") as fh:
        data = fh.read()
    return data, tpl.original_name, template_ref(tpl)


# ── Cycle de vie maker-checker ────────────────────────────────────────────────

def upload_template(file, *, kind: str = KIND_FEUILLE_BESOINS, uploaded_by: str = "") -> FileTemplate:
    """Upload d'un template (maker). `.xlsx` uniquement, ≤ 5 Mo, SHA-256, statut `pending`.

    Le schéma est dérivé dès l'upload (aperçu pour le checker) et re-dérivé, autoritatif,
    à l'activation.
    """
    name = getattr(file, "name", "") or ""
    if not name.lower().endswith(".xlsx"):
        raise TemplateUploadError(
            "EXTENSION_INVALIDE",
            "Format attendu : fichier .xlsx (le .xlsm est refusé — CLAUDE.md §5).",
        )
    size = getattr(file, "size", None)
    if size is not None and size > MAX_TEMPLATE_BYTES:
        raise TemplateUploadError(
            "FICHIER_TROP_VOLUMINEUX",
            "Le fichier dépasse la taille maximale autorisée (5 Mo).",
        )

    last = FileTemplate.objects.filter(kind=kind).order_by("-version").first()
    version = (last.version + 1) if last else 1

    tpl = FileTemplate(
        kind=kind, original_name=name[:255], version=version, uploaded_by=uploaded_by or "",
    )
    tpl.file = file
    tpl.save()

    tpl.sha256 = svc.file_sha256(tpl.file.path)
    try:
        tpl.schema = derive_schema(tpl.file.path)
    except Exception as exc:
        tpl.delete()
        raise TemplateUploadError(
            "CLASSEUR_ILLISIBLE",
            f"Le classeur n'a pas pu être analysé comme template : {exc}",
        )
    tpl.save(update_fields=["sha256", "schema"])

    _audit(actor=uploaded_by, action="dataio.template.upload", tpl=tpl,
           details={"kind": kind, "version": version, "sha256": tpl.sha256})
    return tpl


def activate_template(tpl: FileTemplate, *, activator_sub: str) -> FileTemplate:
    """Activation (checker ≠ maker). Le template actif du même `kind` passe `archived` ;
    le schéma est (re)dérivé du fichier à cet instant — il devient la règle de validation.
    """
    if tpl.status != FileTemplate.Status.PENDING:
        raise TemplateActivationError(
            "STATUT_INVALIDE",
            f"Ce template ne peut pas être activé (statut « {tpl.status} »). "
            f"Seul un template « pending » s'active.",
        )
    if tpl.uploaded_by and tpl.uploaded_by == activator_sub:
        raise TemplateActivationError(
            "MAKER_EGAL_CHECKER",
            "Le même utilisateur ne peut pas téléverser ET activer un template "
            "(principe maker-checker : l'activation doit être faite par un second "
            "administrateur).",
        )

    with transaction.atomic():
        previous = (FileTemplate.objects.select_for_update()
                    .filter(kind=tpl.kind, status=FileTemplate.Status.ACTIVE).first())
        if previous:
            previous.status = FileTemplate.Status.ARCHIVED
            previous.save(update_fields=["status"])
            tpl.supersedes = previous

        tpl.schema = derive_schema(tpl.file.path)
        tpl.status = FileTemplate.Status.ACTIVE
        tpl.activated_by = activator_sub
        tpl.activated_at = timezone.now()
        tpl.save(update_fields=["schema", "status", "activated_by", "activated_at", "supersedes"])

        _audit(actor=activator_sub, action="dataio.template.activate", tpl=tpl,
               details={"kind": tpl.kind, "version": tpl.version,
                        "supersedes": previous.pk if previous else None,
                        "maker": tpl.uploaded_by, "checker": activator_sub})
    return tpl


def _audit(*, actor: str, action: str, tpl: FileTemplate, details: dict) -> None:
    """Journalise l'événement template (l'auditeur doit pouvoir reconstituer qui a activé
    quelle règle de validation, et quand)."""
    from audit.services import record
    record(actor=actor or "", action=action, entity_type="FileTemplate",
           entity_id=str(tpl.pk), details=details)
