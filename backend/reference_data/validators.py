"""
Validation ligne par ligne du fichier Excel «Chaînes de Valeur».

Format attendu (Annexe A du spec) :
  Feuille : value_chains
  Colonnes : code, label, cycle_months, cost_ha_usd, cost_ha_cdf,
             w_semences, w_mecanisation, w_maindoeuvre, w_equipements,
             w_postrecolte, w_logistique, w_commercialisation, w_reserve,
             risk_factor, min_score, base_rate, harvest_months, eligible_guarantees
"""
from __future__ import annotations

from typing import Any

import openpyxl

SHEET_NAME = "value_chains"

MODULE_WEIGHT_COLS = [
    "w_semences", "w_mecanisation", "w_maindoeuvre", "w_equipements",
    "w_postrecolte", "w_logistique", "w_commercialisation", "w_reserve",
]

REQUIRED_COLS = [
    "code", "label", "cycle_months", "cost_ha_usd", "cost_ha_cdf",
    *MODULE_WEIGHT_COLS,
    "risk_factor", "min_score", "base_rate", "harvest_months", "eligible_guarantees",
]

NUMERIC_COLS = [
    "cycle_months", "cost_ha_usd", "cost_ha_cdf",
    *MODULE_WEIGHT_COLS,
    "risk_factor", "min_score", "base_rate",
]

VALID_GUARANTEES = {"epargne", "morale", "foncier", "materiel"}


class ExcelStructureError(Exception):
    """Le fichier Excel ne respecte pas la structure minimale attendue."""


def _to_num(val: Any, col: str, row_idx: int, errors: list[str]) -> float | None:
    if val is None or str(val).strip() == "":
        errors.append(f"Ligne {row_idx} : colonne '{col}' vide ou manquante.")
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        errors.append(
            f"Ligne {row_idx} : colonne '{col}' = «{val}» — valeur numérique attendue."
        )
        return None


def validate_value_chains_sheet(file_path: str) -> dict:
    """
    Parse et valide le fichier Excel chaînes de valeur.
    Retourne {"valid": bool, "errors": [...], "rows": [...]} où rows = données
    parsées (seulement si valid=True, sinon liste partielle des lignes valides).
    """
    errors: list[str] = []
    rows: list[dict] = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise ExcelStructureError(
            f"Impossible d'ouvrir le fichier Excel : {exc}"
        ) from exc

    if SHEET_NAME not in wb.sheetnames:
        raise ExcelStructureError(
            f"Feuille «{SHEET_NAME}» introuvable. "
            f"Feuilles présentes : {', '.join(wb.sheetnames)}. "
            f"Téléchargez le modèle officiel pour obtenir la bonne structure."
        )

    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ExcelStructureError("La feuille «value_chains» est vide.")

    # Header
    header = [
        str(c).strip().lower() if c is not None else ""
        for c in all_rows[0]
    ]
    missing_cols = [col for col in REQUIRED_COLS if col not in header]
    if missing_cols:
        raise ExcelStructureError(
            f"Colonnes requises manquantes : {', '.join(missing_cols)}. "
            f"Colonnes trouvées : {', '.join(c for c in header if c)}."
        )

    col_idx: dict[str, int] = {col: header.index(col) for col in REQUIRED_COLS}
    seen_codes: set[str] = set()

    for row_i, raw_row in enumerate(all_rows[1:], start=2):
        if all(v is None for v in raw_row):
            continue  # ignore blank rows

        def get(col: str) -> Any:
            idx = col_idx.get(col)
            return raw_row[idx] if idx is not None and idx < len(raw_row) else None

        code = str(get("code") or "").strip().upper()
        label = str(get("label") or "").strip()

        row_errors: list[str] = []

        if not code:
            errors.append(f"Ligne {row_i} : colonne 'code' vide — chaque chaîne doit avoir un code unique.")
            continue
        if not label:
            row_errors.append(f"Ligne {row_i} : colonne 'label' vide.")
        if code in seen_codes:
            row_errors.append(f"Ligne {row_i} : code '{code}' en doublon dans le fichier.")
        seen_codes.add(code)

        nums: dict[str, float | None] = {}
        for col in NUMERIC_COLS:
            nums[col] = _to_num(get(col), col, row_i, row_errors)

        # Somme des poids = 100 %
        weight_vals = [nums[c] for c in MODULE_WEIGHT_COLS if nums.get(c) is not None]
        if len(weight_vals) == len(MODULE_WEIGHT_COLS):
            total_w = sum(weight_vals)
            if abs(total_w - 100) > 0.5:
                row_errors.append(
                    f"Ligne {row_i} : la somme des poids modules fait {total_w:.1f} %, attendu 100 %."
                )

        # Taux de base ∈ [1 %, 30 %]
        br = nums.get("base_rate")
        if br is not None and not (1 <= br <= 30):
            row_errors.append(
                f"Ligne {row_i} : taux de base {br} % hors bornes [1 %, 30 %]."
            )

        # risk_factor > 0
        rf = nums.get("risk_factor")
        if rf is not None and rf <= 0:
            row_errors.append(f"Ligne {row_i} : risk_factor doit être > 0 (valeur : {rf}).")

        # min_score ∈ [0, 100]
        ms = nums.get("min_score")
        if ms is not None and not (0 <= ms <= 100):
            row_errors.append(f"Ligne {row_i} : min_score {ms} hors bornes [0, 100].")

        # harvest_months — "3;4" ou "3,4" → [3, 4], chaque élément ∈ [1..12]
        harvest_raw = str(get("harvest_months") or "").strip()
        harvest_months: list[int] = []
        if harvest_raw:
            for part in harvest_raw.replace(",", ";").split(";"):
                part = part.strip()
                if not part:
                    continue
                try:
                    m = int(float(part))
                    if not 1 <= m <= 12:
                        row_errors.append(
                            f"Ligne {row_i} : mois de récolte {m} invalide (attendu 1–12)."
                        )
                    else:
                        harvest_months.append(m)
                except ValueError:
                    row_errors.append(
                        f"Ligne {row_i} : mois de récolte «{part}» non numérique."
                    )

        # eligible_guarantees — "epargne;morale"
        guar_raw = str(get("eligible_guarantees") or "").strip()
        eligible_guarantees = [
            g.strip().lower()
            for g in guar_raw.replace(",", ";").split(";")
            if g.strip()
        ]
        for g in eligible_guarantees:
            if g not in VALID_GUARANTEES:
                row_errors.append(
                    f"Ligne {row_i} : garantie «{g}» inconnue "
                    f"(valeurs acceptées : {', '.join(sorted(VALID_GUARANTEES))})."
                )

        errors.extend(row_errors)

        # N'ajouter la ligne dans rows que si elle est valide
        if not row_errors and all(
            nums.get(c) is not None for c in NUMERIC_COLS
        ):
            module_weights = {
                col.removeprefix("w_"): round(nums[col], 2)
                for col in MODULE_WEIGHT_COLS
            }
            rows.append({
                "code": code,
                "label": label,
                "active": True,
                "cycle_months": int(nums["cycle_months"]),
                "cost_per_hectare_usd": str(round(nums["cost_ha_usd"], 2)),
                "cost_per_hectare_cdf": str(round(nums["cost_ha_cdf"], 2)),
                "module_weights": module_weights,
                "risk_factor": str(round(nums["risk_factor"], 3)),
                "min_score_required": int(nums["min_score"]),
                "base_rate": str(round(nums["base_rate"], 2)),
                "harvest_months": harvest_months,
                "eligible_guarantees": eligible_guarantees,
            })

    return {"valid": len(errors) == 0, "errors": errors, "rows": rows}
