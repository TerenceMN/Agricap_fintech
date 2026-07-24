"""
Ingestion d'un classeur de référence AGRICAP → tables (écriture à l'upload).

Appelé par la commande `import_referentiel` ET par l'endpoint d'upload admin :
une seule logique. Le référentiel v3 étant hétérogène (colonnes différentes par
famille de chaîne), on capture TOUTES les colonnes puis on dérive les concepts
communs par correspondance sémantique d'en-têtes. On n'invente rien : un concept
introuvable reste None (→ `NON ÉVALUABLE` en aval).
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass

import openpyxl
from django.db import transaction

from .chains import BY_SHEET
from .models import InstitutionConfig, ReferenceRange, ReferentielVersion
from .range_parser import to_decimal_range, to_number, to_range

CALIBRATION_SHEET = "16_Calibrage_Gouvernance"


def _norm(text) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFD", str(text))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


@dataclass
class IngestReport:
    version_id: int
    version_label: str
    ranges_by_chain: dict[str, int]
    config_loaded: bool
    warnings: list[str]

    @property
    def total_ranges(self) -> int:
        return sum(self.ranges_by_chain.values())


# --- Détection d'en-tête ---------------------------------------------------
_TITLE_PREFIXES = ("chaine de valeur", "valeurs indicatives", "colonnes", "legende", "note")


def _find_header(ws) -> int | None:
    """Ligne d'en-tête = 1re ligne (1..8) avec ≥4 cellules dont la 1re n'est pas un titre."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        non_empty = [c for c in row if c not in (None, "")]
        first = _norm(row[0]) if row else ""
        if len(non_empty) >= 4 and first and not any(first.startswith(p) for p in _TITLE_PREFIXES):
            return r
    return None


# --- Dérivation sémantique des concepts ------------------------------------
def _headers(ws, header_row) -> list[tuple[int, str, str]]:
    """(index, en-tête brut, en-tête normalisé) pour la ligne d'en-tête."""
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return [(i, str(h).strip(), _norm(h)) for i, h in enumerate(row) if h not in (None, "")]


def _find_range(headers, row, keywords, *, exclude=("median", "mois hauts", "mois bas")):
    """
    Cherche un concept par mots-clés d'en-tête. Si deux colonnes correspondent avec
    « min »/« max », retourne (min, max) ; sinon parse la 1re colonne comme plage.
    Retourne (texte_source, lo, hi).
    """
    matches = [(i, h, n) for (i, h, n) in headers
               if any(k in n for k in keywords) and not any(e in n for e in exclude)]
    if not matches:
        return "", None, None

    def val(i):
        return row[i] if i < len(row) else None

    col_min = next((m for m in matches if "min" in m[2]), None)
    col_max = next((m for m in matches if "max" in m[2]), None)
    if col_min and col_max:
        lo = to_number(val(col_min[0]))
        hi = to_number(val(col_max[0]))
        text = f"{val(col_min[0])}–{val(col_max[0])}"
        if lo is not None and hi is not None and lo > hi:
            lo, hi = hi, lo
        return text, lo, hi

    i, h, n = matches[0]
    raw = val(i)
    lo, hi = to_range(raw)
    return (str(raw) if raw not in (None, "") else ""), lo, hi


_RENDEMENT_KW = ("rendement", "production", "poids vif", "poids moyen", "gmq", "ponte",
                 "prolificite", "densite", "capacite", "litres", "taux de ponte", "parametre technique")
_COUT_KW = ("cout de production", "cout de transfo", "cout matiere", "cout",)
_PRIX_KW = ("prix de vente", "prix unitaire", "prix",)
_PERTE_KW = ("mortalite", "perte", "taux de survie", "survie", "colonisation",)


def _derive_perte(headers, row) -> tuple[str, float | None]:
    """Perte/mortalité en fraction. Une métrique de survie/colonisation est inversée."""
    text, lo, hi = _find_range(headers, row, _PERTE_KW)
    if hi is None:
        return text, None
    # Détecte si le concept trouvé est une « survie/colonisation » (métrique positive).
    is_survie = any(("survie" in n or "colonisation" in n) and any(k in n for k in _PERTE_KW)
                    for (_, _, n) in headers)
    return (text, round(1 - lo, 4) if (is_survie and lo is not None) else hi)


def _build_range(version, chain, headers, row):
    """
    Construit un `ReferenceRange` (non sauvegardé) depuis (headers, row).
    Renvoie ("stop", None) sur une légende de bas de tableau, ("skip", None) sur
    une ligne vide, ("row", ReferenceRange) sinon. Partagé par les deux sources de
    données : le classeur (`_ingest_chain_sheet`) et les lignes éditées en base
    (`rebuild_chain_from_records`), pour une dérivation identique.
    """
    name = row[0] if row else None
    nname = _norm(name)
    if not nname:
        return "skip", None
    if any(nname.startswith(p) for p in _TITLE_PREFIXES):
        return "stop", None  # légende de bas de tableau → fin

    def by_kw(*keywords):
        """Valeur brute de la 1re colonne dont l'en-tête contient un des mots-clés."""
        for (i, h, n) in headers:
            if any(k in n for k in keywords):
                return row[i] if i < len(row) else None
        return None

    cols = {h: (row[i] if i < len(row) else None) for (i, h, n) in headers}
    rend_param = next((h for (i, h, n) in headers if any(k in n for k in _RENDEMENT_KW)), "")
    _, rend_lo, rend_hi = _find_range(headers, row, _RENDEMENT_KW)
    if rend_lo is None and rend_hi is None:
        # Layout élevage : le libellé (« GMQ… ») et ses bornes « Valeur min/max »
        # sont dans des colonnes distinctes.
        _, rend_lo, rend_hi = _find_range(headers, row, ("valeur",))
    cout_txt, cout_lo, cout_hi = _find_range(headers, row, _COUT_KW)
    prix_txt, prix_lo, prix_hi = _find_range(headers, row, _PRIX_KW)
    perte_txt, perte_max = _derive_perte(headers, row)

    return "row", ReferenceRange(
        version=version,
        chain_code=chain.code, chain_slug=chain.slug, chain_libelle=chain.libelle,
        name=str(name).strip()[:160],
        systeme=str(by_kw("systeme") or "").strip()[:160],
        cycle_months=to_number(by_kw("cycle", "duree du cycle")),
        parametre_cle=str(rend_param)[:160],
        unite=str(by_kw("unite") or "").strip()[:60],
        rendement_min=rend_lo, rendement_max=rend_hi,
        cout_text=cout_txt[:120], cout_min=cout_lo, cout_max=cout_hi,
        prix_text=prix_txt[:120], prix_min=prix_lo, prix_max=prix_hi,
        perte_text=perte_txt[:120], perte_max=perte_max,
        statut=str(by_kw("statut") or "").strip()[:60],
        source=str(by_kw("source") or "").strip()[:255],
        date_maj=str(by_kw("date maj") or "").strip()[:40],
        zone=str(by_kw("zone") or "").strip()[:200],
        observations=str(by_kw("observations") or "").strip(),
        columns={str(k): (None if v is None else str(v)) for k, v in cols.items()},
    )


def _ingest_chain_sheet(version, chain, ws) -> tuple[int, list[str]]:
    header_row = _find_header(ws)
    if not header_row:
        return 0, [f"{chain.sheet}: en-tête introuvable, feuille ignorée."]
    headers = _headers(ws, header_row)

    objs: list[ReferenceRange] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        status, rr = _build_range(version, chain, headers, row)
        if status == "stop":
            break
        if status == "skip":
            continue
        objs.append(rr)

    ReferenceRange.objects.bulk_create(objs)
    return len(objs), []


@transaction.atomic
def rebuild_chain_from_records(version, chain, header_names, row_dicts) -> int:
    """
    Re-dérive les `ReferenceRange` d'une chaîne à partir des lignes ÉDITÉES en base
    (couche générique). `header_names` = colonnes ordonnées ; `row_dicts` = valeurs
    par en-tête. Remplace les plages typées de cette chaîne dans `version` — c'est
    ce que le moteur relit après une correction admin.
    """
    headers = [(i, str(nm), _norm(nm)) for i, nm in enumerate(header_names)]
    objs: list[ReferenceRange] = []
    for rd in row_dicts:
        row = tuple(rd.get(nm) for nm in header_names)
        status, rr = _build_range(version, chain, headers, row)
        if status == "stop":
            break
        if status == "skip":
            continue
        objs.append(rr)

    ReferenceRange.objects.filter(version=version, chain_code=chain.code).delete()
    ReferenceRange.objects.bulk_create(objs)
    return len(objs)


@transaction.atomic
def rebuild_config_from_records(version, header_names, row_dicts) -> bool:
    """Re-dérive `InstitutionConfig` (feuille 16) depuis les lignes éditées en base.

    Lecture en `to_decimal_range` (et non `to_range`) : ces cellules deviennent les
    seuils du moteur de décision, pas des ordres de grandeur — elles ne doivent
    jamais transiter par un `float` (principe 4, cf. `InstitutionConfig`).
    """
    if not header_names:
        return False
    cfg = InstitutionConfig(version=version, is_active=True)
    raw: list[list[str]] = []
    found = False
    for rd in row_dicts:
        cells = [rd.get(nm) for nm in header_names]
        non_empty = [str(c) for c in cells if c not in (None, "")]
        if non_empty:
            raw.append(non_empty)
        label = _norm(cells[0]) if cells else ""
        value = cells[1] if len(cells) > 1 else None
        for needle, attr in _CFG_KEYS.items():
            if needle in label:
                lo, hi = to_decimal_range(value)
                if hi is not None:
                    setattr(cfg, attr, hi)
                    found = True
                break
    cfg.raw = {"feuille_16": raw}
    InstitutionConfig.objects.filter(is_active=True).update(is_active=False)
    cfg.save()
    return found


# Paramètre feuille 16 (normalisé) → attribut InstitutionConfig.
_CFG_KEYS = {
    "seuil dscr (avis": "seuil_dscr",
    "seuil dscr stresse": "seuil_dscr_stresse",
    "couverture des garanties": "couverture_min",
    "score global minimal": "score_global_min",
    "ponderation comportemental": "poids_comportemental",
}


def _ingest_calibration(version, ws) -> tuple[bool, list[str]]:
    """Feuille 16 → `InstitutionConfig`. `to_decimal_range` : ces valeurs sont
    les paramètres du moteur, elles restent exactes de la cellule à la colonne."""
    cfg = InstitutionConfig(version=version, is_active=True)
    raw: list[list] = []
    found = False
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c not in (None, "")]
        if cells:
            raw.append([str(c) for c in cells])
        if len(row) < 2:
            continue
        label = _norm(row[0])
        for needle, attr in _CFG_KEYS.items():
            if needle in label:
                lo, hi = to_decimal_range(row[1])
                if hi is not None:
                    setattr(cfg, attr, hi)
                    found = True
                break
    cfg.raw = {"feuille_16": raw}
    InstitutionConfig.objects.filter(is_active=True).update(is_active=False)
    cfg.save()
    return found, ([] if found else ["Feuille 16: aucun seuil reconnu (valeurs par défaut)."])


@transaction.atomic
def ingest_workbook(path_or_file, *, label: str, source_filename: str = "") -> IngestReport:
    """Parse un classeur référentiel et remplit les tables. Renvoie un rapport."""
    wb = openpyxl.load_workbook(path_or_file, data_only=True, read_only=True)
    version = ReferentielVersion.objects.create(label=label, source_filename=source_filename)
    ReferentielVersion.objects.exclude(pk=version.pk).update(is_active=False)

    ranges_by_chain: dict[str, int] = {}
    warnings: list[str] = []
    for sheet_name in wb.sheetnames:
        chain = BY_SHEET.get(sheet_name)
        if chain:
            n, w = _ingest_chain_sheet(version, chain, wb[sheet_name])
            ranges_by_chain[chain.code] = n
            warnings += w

    config_loaded = False
    if CALIBRATION_SHEET in wb.sheetnames:
        config_loaded, w = _ingest_calibration(version, wb[CALIBRATION_SHEET])
        warnings += w
    else:
        warnings.append("Feuille 16 absente : configuration par défaut.")
        InstitutionConfig.objects.filter(is_active=True).update(is_active=False)
        InstitutionConfig.objects.create(version=version, is_active=True)

    wb.close()
    return IngestReport(
        version_id=version.pk, version_label=version.label,
        ranges_by_chain=ranges_by_chain, config_loaded=config_loaded, warnings=warnings,
    )
