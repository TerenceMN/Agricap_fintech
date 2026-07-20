"""
Générateur du modèle Excel «Feuille de Besoins» (GET /api/credits/needs-sheet-template/).
Produit dynamiquement depuis le référentiel filière (ValueChain).
"""
from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Couleurs thématiques
COLOR_HEADER = "2E7D32"   # vert foncé
COLOR_SUBHEADER = "A5D6A7"  # vert clair
COLOR_EXAMPLE = "F1F8E9"  # vert très clair
COLOR_TITLE = "1B5E20"   # vert très foncé

SHEET_DEFINITIONS = [
    {
        "num": 2,
        "title": "Semences & Intrants",
        "module": "semences",
        "columns": [
            ("Désignation de l'intrant", 35),
            ("Unité", 10),
            ("Quantité", 12),
            ("Prix unitaire (USD)", 18),
            ("Total (USD)", 14),
            ("Fournisseur souhaité", 28),
        ],
        "example_rows": [
            ["Semences certifiées", "kg", 50, 8.00, "=C{row}*D{row}", "Kivu Seeds & Co."],
            ["Engrais NPK", "sac 50kg", 10, 45.00, "=C{row}*D{row}", "Agro-Input SARL"],
            ["Pesticide systémique", "L", 5, 22.00, "=C{row}*D{row}", ""],
        ],
    },
    {
        "num": 3,
        "title": "Mécanisation & Équipements",
        "module": "mecanisation",
        "columns": [
            ("Opération / Équipement", 35),
            ("Unité", 10),
            ("Quantité", 12),
            ("Prix unitaire (USD)", 18),
            ("Total (USD)", 14),
            ("Fournisseur souhaité", 28),
        ],
        "example_rows": [
            ["Location tracteur (labour)", "heure", 8, 35.00, "=C{row}*D{row}", "Méca-Kivu"],
            ["Houes & outils manuels", "unité", 15, 12.00, "=C{row}*D{row}", ""],
        ],
    },
    {
        "num": 4,
        "title": "Main-d'œuvre",
        "module": "maindoeuvre",
        "columns": [
            ("Tâche / Activité", 35),
            ("Nb personnes", 14),
            ("Nb jours", 10),
            ("Coût/jour (USD)", 14),
            ("Total (USD)", 14),
            ("Prestataire", 25),
        ],
        "example_rows": [
            ["Semis & plantation", 5, 10, 8.00, "=B{row}*C{row}*D{row}", ""],
            ["Désherbage (1er passage)", 8, 6, 8.00, "=B{row}*C{row}*D{row}", ""],
            ["Récolte manuelle", 12, 5, 10.00, "=B{row}*C{row}*D{row}", ""],
        ],
    },
    {
        "num": 5,
        "title": "Récolte & Post-récolte",
        "module": "postrecolte",
        "columns": [
            ("Opération / Équipement", 35),
            ("Unité", 10),
            ("Quantité", 12),
            ("Prix unitaire (USD)", 18),
            ("Total (USD)", 14),
            ("Fournisseur souhaité", 28),
        ],
        "example_rows": [
            ["Sacs de collecte (grains)", "sac", 100, 1.50, "=C{row}*D{row}", ""],
            ["Location dépulpeuse", "jour", 3, 40.00, "=C{row}*D{row}", "Coopérative Locale"],
            ["Séchage solaire (nattes)", "m²", 50, 2.00, "=C{row}*D{row}", ""],
        ],
    },
    {
        "num": 6,
        "title": "Logistique & Commercialisation",
        "module": "logistique",
        "columns": [
            ("Description", 35),
            ("Unité", 10),
            ("Quantité", 12),
            ("Prix unitaire (USD)", 18),
            ("Total (USD)", 14),
            ("Prestataire / Acheteur", 28),
        ],
        "example_rows": [
            ["Transport récolte → marché", "voyage", 4, 80.00, "=C{row}*D{row}", "Trans-Kivu"],
            ["Commission agent commercial", "% valeur", 1, 250.00, "=C{row}*D{row}", ""],
        ],
    },
    {
        "num": 7,
        "title": "Réserve d'exploitation",
        "module": "reserve",
        "columns": [
            ("Poste de dépense", 35),
            ("Unité", 10),
            ("Quantité", 12),
            ("Prix unitaire (USD)", 18),
            ("Total (USD)", 14),
            ("Remarques", 28),
        ],
        "example_rows": [
            ["Réserve imprévus (5 % total)", "forfait", 1, 0.00, "=C{row}*D{row}", "À calculer"],
            ["Frais bancaires & transferts", "forfait", 1, 50.00, "=C{row}*D{row}", ""],
        ],
    },
]


def _header_style(ws, row: int, col_count: int, title: str) -> None:
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True, color="FFFFFF", size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20


def _col_header_style(ws, row: int, col_count: int) -> None:
    fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    font = Font(bold=True, size=10)
    thin = Side(border_style="thin", color="2E7D32")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = Border(bottom=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _instruction_note(ws, row: int, col_count: int, text: str) -> None:
    fill = PatternFill("solid", fgColor="FFF9C4")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill
    cell.font = Font(italic=True, color="555500", size=9)
    cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 28


def _add_data_sheet(wb: openpyxl.Workbook, sheet_def: dict, value_chain=None) -> None:
    """Ajoute une feuille de données (feuilles 2-7)."""
    ws = wb.create_sheet(title=sheet_def["title"])
    columns = sheet_def["columns"]
    n_cols = len(columns)
    current_row = 1

    # ── En-tête feuille ──
    _header_style(ws, current_row, n_cols, f"Section {sheet_def['num']} — {sheet_def['title']}")
    current_row += 1

    # Note d'instruction
    module = sheet_def["module"]
    std_amount = ""
    if value_chain:
        pct = (value_chain.module_weights or {}).get(module, 0)
        if pct:
            ha_cost = float(value_chain.cost_per_hectare_usd)
            std_amount = f" (standard référentiel : {pct:.0f} % du budget, soit ~{ha_cost * pct / 100:.0f} USD/ha)"
    _instruction_note(
        ws, current_row, n_cols,
        f"ℹ️  Listez tous vos besoins pour le module «{module}»{std_amount}. "
        f"Les totaux sont recalculés automatiquement côté serveur.",
    )
    current_row += 1

    # En-têtes colonnes
    _col_header_style(ws, current_row, n_cols)
    for c, (col_name, col_width) in enumerate(columns, start=1):
        ws.cell(row=current_row, column=c, value=col_name)
        ws.column_dimensions[get_column_letter(c)].width = col_width
    current_row += 1

    # Lignes d'exemple
    fill_example = PatternFill("solid", fgColor=COLOR_EXAMPLE)
    for example in sheet_def["example_rows"]:
        for c, val in enumerate(example, start=1):
            cell = ws.cell(row=current_row, column=c)
            if isinstance(val, str) and val.startswith("="):
                # Formule Excel avec substitution du numéro de ligne
                cell.value = val.replace("{row}", str(current_row))
            else:
                cell.value = val
            cell.fill = fill_example
        current_row += 1

    # Lignes vides pour saisie
    for _ in range(10):
        thin = Side(border_style="thin", color="CCCCCC")
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=current_row, column=c)
            cell.border = Border(bottom=thin, right=thin)
        current_row += 1

    # Ligne TOTAL
    total_col_idx = next(
        (i + 1 for i, (name, _) in enumerate(columns) if "total" in name.lower()),
        None,
    )
    if total_col_idx:
        label_col_idx = 1 if n_cols >= 1 else None
        ws.cell(row=current_row, column=1, value="TOTAL").font = Font(bold=True)
        total_start = current_row - 10 - len(sheet_def["example_rows"])
        total_end = current_row - 1
        total_cell = ws.cell(row=current_row, column=total_col_idx)
        total_cell.value = (
            f"=SUM({get_column_letter(total_col_idx)}{total_start}:"
            f"{get_column_letter(total_col_idx)}{total_end})"
        )
        total_cell.font = Font(bold=True)
        fill_total = PatternFill("solid", fgColor="E8F5E9")
        for c in range(1, n_cols + 1):
            ws.cell(row=current_row, column=c).fill = fill_total


def generate_needs_sheet_template(value_chain=None, applicant_name: str = "") -> bytes:
    """
    Génère le classeur Excel modèle Feuille de Besoins.
    Retourne les bytes du fichier .xlsx.
    """
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Identification ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Identification"

    fill_title = PatternFill("solid", fgColor=COLOR_TITLE)
    ws1.merge_cells("A1:F1")
    cell = ws1["A1"]
    cell.value = "FEUILLE DE BESOINS — DEMANDE DE CRÉDIT AGRICOLE"
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = fill_title
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    fill_sub = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    ws1.merge_cells("A2:F2")
    sub = ws1["A2"]
    vc_label = value_chain.label if value_chain else "[Sélectionner la chaîne de valeur]"
    sub.value = f"Filière : {vc_label} | Modèle généré par AGRICAP"
    sub.font = Font(italic=True, size=10)
    sub.fill = fill_sub
    sub.alignment = Alignment(horizontal="center")

    fields = [
        ("Nom / Coopérative :", applicant_name or "[Votre nom ou raison sociale]"),
        ("Superficie concernée (ha) :", ""),
        ("Localisation :", ""),
        ("Date de la demande :", ""),
        ("Chaîne de valeur :", vc_label),
        ("Devise :", "USD"),
    ]
    for i, (label, default) in enumerate(fields, start=4):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=default)
        ws1.merge_cells(f"B{i}:F{i}")

    ws1.merge_cells("A11:F11")
    note = ws1["A11"]
    note.value = (
        "⚠️  INSTRUCTIONS : Complétez les feuilles 2 à 7. "
        "N'ajoutez pas de colonnes. Les totaux sont recalculés par le système."
    )
    note.font = Font(italic=True, color="C62828", size=9)
    note.fill = PatternFill("solid", fgColor="FFEBEE")
    note.alignment = Alignment(wrap_text=True)

    # Si référentiel disponible : tableau de référence
    if value_chain:
        ws1["A13"] = "Budget de référence filière (pour information — à ajuster selon vos besoins)"
        ws1["A13"].font = Font(bold=True, size=10)
        ws1.merge_cells("A13:F13")
        ws1["A14"] = "Module"
        ws1["B14"] = "Poids (%)"
        ws1["C14"] = f"Budget/ha ({value_chain.cost_per_hectare_usd} USD/ha)"
        for c in range(1, 4):
            ws1.cell(14, c).fill = fill_sub
            ws1.cell(14, c).font = Font(bold=True)
        row_ref = 15
        for mod, pct in (value_chain.module_weights or {}).items():
            ws1.cell(row_ref, 1).value = mod
            ws1.cell(row_ref, 2).value = f"{pct:.0f} %"
            ws1.cell(row_ref, 3).value = round(float(value_chain.cost_per_hectare_usd) * pct / 100, 2)
            row_ref += 1

    ws1.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws1.column_dimensions[col].width = 18

    # ── Feuilles 2-7 ────────────────────────────────────────────────────────
    for sheet_def in SHEET_DEFINITIONS:
        _add_data_sheet(wb, sheet_def, value_chain=value_chain)

    # Sauvegarder en mémoire
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
