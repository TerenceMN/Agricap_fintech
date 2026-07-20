"""Tests du lot 2 — la feuille de besoins devient un DataSource.

Couverture :
  - les 6 contrôles de validation de la SPEC, avec leur code d'erreur
  - ingestion dataio : kind, sha256, rattachement au dossier, versionnage
  - `extract_module_totals` lit les DataRecord — le fichier peut disparaître
  - invariant : Σ des 8 modules = TOTAL GÉNÉRAL du classeur
  - suppression refusée d'une pièce probante (409)
"""
from __future__ import annotations

import io
import os
import shutil
import tempfile
from decimal import Decimal

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

from credits.models import CreditApplication
from credits.needs_sheet import (
    MODULE_CODES,
    NeedsSheetValidationError,
    extract_module_totals,
    parse_and_ingest,
    validate_needs_sheet,
)
from dataio.models import DataSource, KIND_FEUILLE_BESOINS, STATUS_COMMITTED
from dataio.services import SourceProtected, delete_source

HEADER_4 = [
    "N°", "Rubrique", "Description détaillée", "Unité", "Quantité", "Coût unitaire",
    "Fréquence", "Montant total", "Période du cycle", "Financement souhaité",
    "Observations", "Validation analyste (réservé)",
]
HEADER_5 = ["Rubrique", "Total rubrique", "Part du total", "Commentaire"]

#: Libellés officiels des 8 rubriques, dans l'ordre de la feuille 5 du modèle.
RUBRIQUES = [
    ("Semences & Intrants", "semences"),
    ("Opérations mécanisées", "mecanisation"),
    ("Main d'œuvre", "maindoeuvre"),
    ("Équipement & petit matériel", "equipements"),
    ("Récolte & post-récolte", "postrecolte"),
    ("Logistique", "logistique"),
    ("Commercialisation", "commercialisation"),
    ("Réserve d'exploitation", "reserve"),
]

#: Cas de référence : le classeur modèle AGRICAP (1 330 USD, cf. SPEC annexe).
LIGNES_MODELE = [
    ("Semences & Intrants", "Semences améliorées", "kg", 50, 12, 1, 600),
    ("Opérations mécanisées", "Labour mécanisé", "ha", 25, 18, 1, 450),
    ("Main d'œuvre", "Désherbage", "j/h", 20, 7, 2, 280),
]


def build_workbook(
    lignes=None,
    *,
    totaux5: dict[str, float] | None = None,
    grand_total: float | None = None,
    header4: list | None = None,
    header5: list | None = None,
    omit_sheets: tuple[str, ...] = (),
    omit_rubriques: tuple[str, ...] = (),
    extra_sheets: int = 0,
) -> bytes:
    """Construit un classeur de feuille de besoins et renvoie ses octets.

    Par défaut la feuille 5 est calculée depuis la feuille 4 (classeur cohérent).
    Les paramètres permettent de fabriquer chaque cas d'erreur de la SPEC.
    """
    lignes = LIGNES_MODELE if lignes is None else lignes

    if totaux5 is None:
        totaux5 = {code: 0.0 for _lbl, code in RUBRIQUES}
        for rubrique, *_rest, total in lignes:
            code = next((c for lbl, c in RUBRIQUES if lbl == rubrique), None)
            if code and isinstance(total, (int, float)):
                totaux5[code] = totaux5.get(code, 0.0) + total
    if grand_total is None:
        grand_total = sum(totaux5.values())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if "4_Besoins_Financiers" not in omit_sheets:
        ws4 = wb.create_sheet("4_Besoins_Financiers")
        ws4.append(["4. Besoins financiers détaillés du cycle"])
        ws4.append(["Saisir uniquement les cellules jaunes."])
        ws4.append([])
        ws4.append(header4 if header4 is not None else HEADER_4)
        for i, (rubrique, designation, unite, qte, cout, freq, total) in enumerate(lignes, 1):
            ws4.append([i, rubrique, designation, unite, qte, cout, freq, total,
                        "Préparation", "Crédit", "", None])

    if "5_Synthese_Besoins" not in omit_sheets:
        ws5 = wb.create_sheet("5_Synthese_Besoins")
        ws5.append(["5. Synthèse des besoins financiers"])
        ws5.append(["Feuille calculée automatiquement — ne rien saisir."])
        ws5.append(header5 if header5 is not None else HEADER_5)
        for label, code in RUBRIQUES:
            if label in omit_rubriques:
                continue
            ws5.append([label, totaux5.get(code, 0.0), 0, None])
        ws5.append(["TOTAL GÉNÉRAL", grand_total, 1, None])

    for n in range(extra_sheets):
        wb.create_sheet(f"Annexe_{n}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(data: bytes, name: str = "feuille_besoins.xlsx") -> SimpleUploadedFile:
    return SimpleUploadedFile(
        name, data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _write_tmp(data: bytes) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(data)
    tmp.close()
    return tmp.name


def _codes(errors: list[dict]) -> set[str]:
    return {e["code"] for e in errors}


class ValidationTests(TestCase):
    """Les 6 contrôles du tableau de la SPEC, un par un."""

    def _validate(self, **kwargs) -> list[dict]:
        path = _write_tmp(build_workbook(**kwargs))
        try:
            return validate_needs_sheet(path)
        finally:
            os.unlink(path)

    def test_classeur_conforme_ne_produit_aucune_erreur(self):
        self.assertEqual(self._validate(), [])

    def test_feuille_manquante(self):
        errors = self._validate(omit_sheets=("5_Synthese_Besoins",))
        self.assertEqual(_codes(errors), {"FEUILLE_MANQUANTE"})

    def test_colonne_manquante(self):
        header = list(HEADER_4)
        header[4] = "Nb"          # « Quantité » n'existe plus sous un nom reconnu
        errors = self._validate(header4=header)
        self.assertEqual(_codes(errors), {"COLONNE_MANQUANTE"})

    def test_rubrique_manquante(self):
        errors = self._validate(omit_rubriques=("Logistique", "Commercialisation"))
        self.assertEqual(_codes(errors), {"RUBRIQUE_MANQUANTE"})
        self.assertEqual(len(errors), 2)

    def test_type_invalide_sur_texte_et_sur_negatif(self):
        lignes = [
            ("Semences & Intrants", "Semences", "kg", "beaucoup", 12, 1, 600),
            ("Main d'œuvre", "Désherbage", "j/h", 20, -7, 2, 280),
        ]
        errors = self._validate(lignes=lignes)
        self.assertEqual(_codes(errors), {"TYPE_INVALIDE"})
        self.assertEqual(len(errors), 2)

    def test_incoherence_interne_detecte_une_feuille_5_ecrasee_a_la_main(self):
        """Le contrôle central : feuille 5 retouchée sans passer par la feuille 4."""
        totaux = {code: 0.0 for _l, code in RUBRIQUES}
        totaux.update({"semences": 810.0, "mecanisation": 450.0, "maindoeuvre": 280.0})
        errors = self._validate(totaux5=totaux, grand_total=1540.0)
        self.assertEqual(_codes(errors), {"INCOHERENCE_INTERNE"})
        self.assertEqual(len(errors), 1)
        message = errors[0]["message"]
        self.assertIn("810", message)   # ce qu'annonce la feuille 5
        self.assertIn("600", message)   # ce que dit réellement la feuille 4

    def test_ecart_sous_le_centime_reste_accepte(self):
        totaux = {code: 0.0 for _l, code in RUBRIQUES}
        totaux.update({"semences": 600.004, "mecanisation": 450.0, "maindoeuvre": 280.0})
        self.assertEqual(self._validate(totaux5=totaux, grand_total=1330.0), [])

    def test_total_incoherent(self):
        errors = self._validate(grand_total=1500.0)
        self.assertEqual(_codes(errors), {"TOTAL_INCOHERENT"})

    def test_rubrique_inconnue_en_feuille_4(self):
        lignes = LIGNES_MODELE + [("Divers imprévus", "Bricoles", "u", 1, 10, 1, 10)]
        errors = self._validate(lignes=lignes)
        self.assertEqual(_codes(errors), {"RUBRIQUE_INCONNUE"})

    def test_ligne_de_sous_total_de_la_feuille_4_est_ignoree(self):
        """Le modèle officiel porte une ligne « TOTAL BESOINS DU CYCLE » en feuille 4 :
        la compter doublerait la rubrique et ferait tomber INCOHERENCE_INTERNE à tort."""
        lignes = LIGNES_MODELE + [
            ("TOTAL BESOINS DU CYCLE", "", "", None, None, None, 1330),
        ]
        self.assertEqual(self._validate(lignes=lignes), [])

    def test_le_modele_officiel_livre_aux_clients_passe_la_validation(self):
        """Principe 11 : le fichier téléchargé doit être celui contre lequel on valide."""
        path = os.path.join(
            os.path.dirname(__file__), "static", "credits", "feuille_besoins_template.xlsx",
        )
        if not os.path.exists(path):       # pragma: no cover - dépend du dépôt
            self.skipTest("Modèle statique absent du dépôt.")
        self.assertEqual(validate_needs_sheet(path), [])

    def test_arret_au_premier_etage_en_echec(self):
        """Une feuille absente ne fait pas remonter d'erreurs de cohérence."""
        errors = self._validate(omit_sheets=("4_Besoins_Financiers",), grand_total=99.0)
        self.assertEqual(_codes(errors), {"FEUILLE_MANQUANTE"})


class IngestionTests(TestCase):
    """Ingestion dataio : rattachement, versionnage, lecture depuis les tables."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._media = tempfile.mkdtemp(prefix="agricap-test-media-")
        cls._override = override_settings(MEDIA_ROOT=cls._media)
        cls._override.enable()

    @classmethod
    def tearDownClass(cls):
        cls._override.disable()
        shutil.rmtree(cls._media, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        from accounts.models import FintechUser
        self.client_user, _ = FintechUser.objects.get_or_create(
            sub="sub-client-fb", defaults={"full_name": "Client FB", "phone": "+243000000001"},
        )
        self.app = CreditApplication.objects.create(
            code="CRED-TEST-FB-0001",
            client=self.client_user,
            amount_requested=Decimal("1050"),
            currency="USD",
            status=CreditApplication.Status.DRAFT,
        )

    def _ingest(self, **kwargs) -> dict:
        return parse_and_ingest(
            _upload(build_workbook(**kwargs)), self.app, uploaded_by="sub-client-fb",
        )

    def test_ingestion_nominale(self):
        result = self._ingest()
        self.app.refresh_from_db()
        source = self.app.needs_source

        self.assertIsNotNone(source)
        self.assertEqual(source.pk, result["needs_source_id"])
        self.assertEqual(source.kind, KIND_FEUILLE_BESOINS)
        self.assertEqual(source.status, STATUS_COMMITTED)
        self.assertTrue(source.is_current)
        self.assertEqual(source.credit_application_id, self.app.pk)
        self.assertEqual(source.dataset_key, "fb__CRED-TEST-FB-0001")
        self.assertEqual(len(source.sha256), 64)
        # Seules les deux feuilles de données sont ingérées.
        self.assertEqual(
            sorted(t.name for t in source.tables.all()),
            ["4_Besoins_Financiers", "5_Synthese_Besoins"],
        )

    def test_totaux_par_module_et_invariant_de_somme(self):
        result = self._ingest()
        totals = {k: Decimal(v) for k, v in result["totals"].items()}

        self.assertEqual(set(totals), set(MODULE_CODES))
        self.assertEqual(totals["semences"], Decimal("600.00"))
        self.assertEqual(totals["mecanisation"], Decimal("450.00"))
        self.assertEqual(totals["maindoeuvre"], Decimal("280.00"))
        self.assertEqual(totals["reserve"], Decimal("0.00"))
        # Invariant : Σ des 8 modules = TOTAL GÉNÉRAL du classeur.
        self.assertEqual(sum(totals.values()), Decimal("1330.00"))
        self.assertEqual(Decimal(result["grand_total"]), Decimal("1330.00"))

    def test_extraction_lit_les_datarecord_pas_le_fichier(self):
        """Le fichier peut disparaître : les totaux restent ceux de la base."""
        self._ingest()
        self.app.refresh_from_db()
        source = self.app.needs_source

        os.unlink(source.file.path)
        totals = extract_module_totals(source)

        self.assertEqual(totals["semences"], Decimal("600.00"))
        self.assertEqual(sum(totals.values()), Decimal("1330.00"))

    def test_reupload_cree_une_nouvelle_revision_et_conserve_l_historique(self):
        first = self._ingest()
        lignes = [("Semences & Intrants", "Semences", "kg", 50, 20, 1, 1000)]
        second = self._ingest(lignes=lignes)

        self.assertEqual(first["revision"], 1)
        self.assertEqual(second["revision"], 2)
        self.assertNotEqual(first["sha256"], second["sha256"])

        old = DataSource.objects.get(pk=first["needs_source_id"])
        new = DataSource.objects.get(pk=second["needs_source_id"])
        self.assertFalse(old.is_current)
        self.assertTrue(new.is_current)
        self.assertEqual(new.supersedes_id, old.pk)

        self.app.refresh_from_db()
        self.assertEqual(self.app.needs_source_id, new.pk)
        # L'ancienne révision reste consultable ligne à ligne.
        self.assertEqual(old.tables.count(), 2)

    def test_classeur_invalide_leve_une_erreur_structuree(self):
        with self.assertRaises(NeedsSheetValidationError) as ctx:
            self._ingest(grand_total=9999.0)

        errors = ctx.exception.errors
        self.assertEqual(_codes(errors), {"TOTAL_INCOHERENT"})
        self.assertTrue(all({"code", "message"} <= set(e) for e in errors))
        # Rien n'est ingéré : le dossier n'a pas de source courante.
        self.app.refresh_from_db()
        self.assertIsNone(self.app.needs_source_id)

    def test_classeur_trop_gros_n_est_pas_une_feuille_de_besoins(self):
        with self.assertRaises(NeedsSheetValidationError) as ctx:
            self._ingest(extra_sheets=12)
        self.assertEqual(_codes(ctx.exception.errors), {"CLASSEUR_NON_RECONNU"})

    def test_suppression_refusee_quand_le_dossier_est_soumis(self):
        self._ingest()
        self.app.refresh_from_db()
        source = self.app.needs_source

        self.app.status = CreditApplication.Status.SUBMITTED
        self.app.save(update_fields=["status"])

        with self.assertRaises(SourceProtected) as ctx:
            delete_source(DataSource.objects.get(pk=source.pk))
        self.assertEqual(ctx.exception.code, "NEEDS_SOURCE_PROTECTED")
        self.assertTrue(DataSource.objects.filter(pk=source.pk).exists())

    def test_suppression_refusee_sur_la_revision_courante_d_un_brouillon(self):
        self._ingest()
        self.app.refresh_from_db()
        with self.assertRaises(SourceProtected) as ctx:
            delete_source(DataSource.objects.get(pk=self.app.needs_source_id))
        self.assertEqual(ctx.exception.code, "NEEDS_SOURCE_IN_USE")

    def test_lignage_du_scoring_permet_de_rejouer_l_analyse(self):
        """Le rapport doit porter needs_source_id + revision + sha256 (principe 1)."""
        from credits.views import _load_needs_totals

        self._ingest()
        self.app.refresh_from_db()
        totals, lineage = _load_needs_totals(self.app)

        self.assertEqual(totals["semences"], 600.0)
        self.assertEqual(lineage["needs_source_id"], self.app.needs_source_id)
        self.assertEqual(lineage["revision"], 1)
        self.assertEqual(len(lineage["sha256"]), 64)
        self.assertEqual(lineage["dataset_key"], "fb__CRED-TEST-FB-0001")

    def test_sans_feuille_ingeree_aucun_total_n_est_inventé(self):
        from credits.views import _load_needs_totals
        totals, lineage = _load_needs_totals(self.app)
        self.assertIsNone(totals)
        self.assertEqual(lineage, {})

    def test_le_client_ne_voit_pas_le_lignage_ni_le_bareme(self):
        """Anti-gaming (principe 7) : le serializer client reste une liste blanche."""
        from credits.view_context import ViewContextService

        self._ingest()
        self.app.refresh_from_db()
        self.app.score_result = {
            "score": 71.0, "eligible": True,
            "breakdown": [{"code": "needs_coherence", "weight": 1.2, "maxPoints": 25}],
            "needsSource": {"sha256": "x" * 64},
            "needsTotals": {"semences": 600.0},
        }
        self.app.save(update_fields=["score_result"])

        data = ViewContextService(sub="sub-client-fb", roles=["client"]).serialize_for_role(self.app)
        # `scoreResult` est soit masqué en entier, soit réduit à une liste blanche —
        # dans les deux cas ni le barème ni le lignage ne fuient.
        score = data.get("scoreResult") or {}
        for leaked in ("breakdown", "needsSource", "needsTotals"):
            self.assertNotIn(leaked, score)

        staff = ViewContextService(sub="sub-agent", roles=["gest_credit"]).serialize_for_role(self.app)
        self.assertEqual((staff.get("scoreResult") or {}).get("score"), 71.0)
