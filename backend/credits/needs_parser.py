"""
Parseur de la Feuille de Besoins (classeur Excel, feuilles 2 à 7).

Fonctionnement :
  1. Reconnaître le fichier (présence des feuilles 2-7 ou noms équivalents)
  2. Pour chaque feuille, détecter les colonnes par synonymes
  3. Extraire les lignes, recalculer les totaux serveur-side
  4. Contrôles de cohérence vs référentiel filière (±30 %)
  5. Vérification des fournisseurs vs annuaire agréé
"""
from __future__ import annotations

import decimal
from typing import Any

import openpyxl

# ── Mapping rubrique (Synthèse) → code module Python ─────────────────────────
_SYNTH_RUBRIQUE_MAP: list[tuple[str, str]] = [
    ("semences",          "semences"),
    ("intrants",          "semences"),
    ("mécanis",           "mecanisation"),
    ("mecanis",           "mecanisation"),
    ("mécani",            "mecanisation"),
    ("main",              "maindoeuvre"),
    ("équipement",        "equipements"),
    ("equipement",        "equipements"),
    ("matériel",          "equipements"),
    ("récolte",           "postrecolte"),
    ("recolte",           "postrecolte"),
    ("post",              "postrecolte"),
    ("logistique",        "logistique"),
    ("commercialisation", "commercialisation"),
    ("réserve",           "reserve"),
    ("reserve",           "reserve"),
]


def _rubrique_to_module(rubrique: str) -> str | None:
    import unicodedata
    nfkd = unicodedata.normalize("NFD", rubrique.lower())
    low = "".join(c for c in nfkd if not unicodedata.combining(c))
    for fragment, code in _SYNTH_RUBRIQUE_MAP:
        if fragment in low:
            return code
    return None


def _find_synthese_sheet(wb):
    """Trouve la feuille Synthèse des Besoins par nom (insensible à la casse/accents)."""
    import unicodedata

    def norm(s):
        nfkd = unicodedata.normalize("NFD", s.lower())
        return "".join(c for c in nfkd if not unicodedata.combining(c))

    for ws in wb.worksheets:
        n = norm(ws.title)
        if "synth" in n and ("besoin" in n or "financier" in n or "5" in n):
            return ws
        if "5_synth" in n or n in ("5_synthese_besoins", "synthese des besoins"):
            return ws
    # Fallback : feuille dont le nom contient "synth"
    for ws in wb.worksheets:
        if "synth" in norm(ws.title):
            return ws
    return None


def _read_synthese_sheet(wb) -> tuple[dict[str, float], float] | tuple[None, None]:
    """
    Lit la feuille Synthèse des Besoins Financiers.

    Retourne (total_by_module, grand_total) ou (None, None) si la feuille
    est absente ou ne contient pas les colonnes attendues.

    Structure attendue :
      en-tête : Rubrique | Total rubrique | Part du total | …
      lignes  : une rubrique par ligne
      dernière ligne significative : TOTAL GÉNÉRAL | <montant> | …
    """
    ws = _find_synthese_sheet(wb)
    if ws is None:
        return None, None

    all_rows = list(ws.iter_rows(values_only=True))

    # Trouver la ligne d'en-tête
    rubrique_col = total_col = None
    header_idx = None
    for i, row in enumerate(all_rows):
        cells = [str(c or "").strip().lower() for c in row]
        # Chercher "rubrique" et "total" dans la même ligne
        rub_i = next((j for j, c in enumerate(cells) if "rubrique" in c), None)
        tot_i = next((j for j, c in enumerate(cells) if "total" in c and "part" not in c), None)
        if rub_i is not None and tot_i is not None:
            rubrique_col = rub_i
            total_col = tot_i
            header_idx = i
            break

    if header_idx is None or rubrique_col is None or total_col is None:
        return None, None

    total_by_module: dict[str, float] = {}
    grand_total: float = 0.0
    grand_total_found = False

    for row in all_rows[header_idx + 1:]:
        if rubrique_col >= len(row) or total_col >= len(row):
            continue
        rubrique_raw = row[rubrique_col]
        total_raw = row[total_col]
        if rubrique_raw is None:
            continue
        rubrique = str(rubrique_raw).strip()
        if not rubrique:
            continue

        val = _to_decimal(total_raw)
        val_f = float(val) if val is not None else 0.0

        import unicodedata
        def norm(s):
            nfkd = unicodedata.normalize("NFD", s.lower())
            return "".join(c for c in nfkd if not unicodedata.combining(c))

        # Ligne TOTAL GÉNÉRAL
        rub_norm = norm(rubrique)
        if "total" in rub_norm and ("general" in rub_norm or "général" in rub_norm or rub_norm.startswith("total")):
            if not grand_total_found:
                grand_total = val_f
                grand_total_found = True
            continue

        # Ligne de module
        mod = _rubrique_to_module(rubrique)
        if mod:
            total_by_module[mod] = total_by_module.get(mod, 0.0) + val_f

    # Si TOTAL GÉNÉRAL absent, calculer depuis les modules
    if not grand_total_found:
        grand_total = sum(total_by_module.values())

    return total_by_module, grand_total


# ── Mapping feuille → module(s) principal(aux) ────────────────────────────────
SHEET_MODULE_MAP = {
    2: ["semences"],
    3: ["mecanisation", "equipements"],
    4: ["maindoeuvre"],
    5: ["postrecolte"],
    6: ["logistique", "commercialisation"],
    7: ["reserve"],
}

# Noms de feuille acceptés (synonymes) — recherche insensible à la casse
SHEET_NAME_SYNONYMS = {
    2: {"semences", "intrants", "semences & intrants", "semences et intrants", "feuille 2", "2"},
    3: {"mecanisation", "mécanisation", "equipements", "équipements", "materiel",
        "mecanique", "feuille 3", "3"},
    4: {"main-d'oeuvre", "main d oeuvre", "maindoeuvre", "main-d'œuvre", "ressources humaines",
        "main d'oeuvre", "feuille 4", "4"},
    5: {"recolte", "récolte", "post-recolte", "post-récolte", "postrecolte", "postrécolte",
        "recolte et post-recolte", "feuille 5", "5"},
    6: {"logistique", "commercialisation", "logistique et commercialisation",
        "transport", "vente", "feuille 6", "6"},
    7: {"reserve", "réserve", "divers", "reserve exploitation", "réserve d'exploitation",
        "feuille 7", "7"},
}

# ── Synonymes de colonnes ────────────────────────────────────────────────────
_ITEM_SYN = {
    "item", "intrant", "désignation", "designation", "description", "libellé", "libelle",
    "rubrique", "poste", "opération", "operation", "tâche", "tache", "produit",
    "équipement", "equipement", "article", "activité", "activite",
}
_QTY_SYN = {
    "quantité", "quantite", "qté", "qte", "nombre", "nb", "qty", "volume", "nbre",
    "nombre de personnes", "nb personnes", "effectif", "personnes",
}
_DAYS_SYN = {"jours", "j/h", "jours/homme", "nb jours", "durée", "duree", "heures", "h"}
_PRICE_SYN = {
    "prix unitaire", "prix/u", "coût unitaire", "cout unitaire", "pu", "prix", "coût",
    "cout", "tarif", "coût/jour", "cout/jour", "prix/jour", "taux journalier",
}
_TOTAL_SYN = {
    "total", "montant total", "montant", "coût total", "cout total",
    "sous-total", "sous total", "total (usd)", "total (cdf)", "total usd", "total cdf",
    "sous-total (usd)", "total général",
}
_SUPPLIER_SYN = {
    "fournisseur", "fournisseur souhaité", "fournisseur souhaite",
    "prestataire", "vendeur", "source", "agrée", "agréé",
}
_UNIT_SYN = {"unité", "unite", "u", "um", "uom"}
_OBS_SYN = {
    "observations", "observation", "obs", "commentaire", "commentaires",
    "remarque", "remarques", "note", "notes",
}


def _norm(s: str) -> str:
    return str(s or "").strip().lower()


def _find_col(header_cells: list[str], synonyms: set[str]) -> int | None:
    for i, cell in enumerate(header_cells):
        if _norm(cell) in synonyms:
            return i
    # Substring match (less strict)
    for i, cell in enumerate(header_cells):
        n = _norm(cell)
        if any(syn in n or n in syn for syn in synonyms if len(syn) > 3):
            return i
    return None


def _to_decimal(val: Any) -> decimal.Decimal | None:
    if val is None:
        return None
    s = str(val).strip().replace(" ", "").replace(",", ".")
    try:
        return decimal.Decimal(s)
    except (decimal.InvalidOperation, ValueError):
        return None


def _find_sheet(wb: openpyxl.Workbook, sheet_num: int) -> openpyxl.worksheet.worksheet.Worksheet | None:
    """Trouve la feuille par position (index 1-based) ou par synonymes de nom."""
    # Position directe (feuille 2 = index 1)
    idx = sheet_num - 1
    if idx < len(wb.worksheets):
        return wb.worksheets[idx]
    # Recherche par nom
    synonyms = SHEET_NAME_SYNONYMS.get(sheet_num, set())
    for ws in wb.worksheets:
        if _norm(ws.title) in synonyms:
            return ws
    return None


def _parse_sheet(ws, sheet_num: int) -> tuple[list[dict], list[str]]:
    """
    Parse une feuille et retourne (items, warnings).
    Détecte la ligne d'en-tête, extrait les données, recalcule les totaux.
    """
    items: list[dict] = []
    warnings: list[str] = []
    modules = SHEET_MODULE_MAP.get(sheet_num, ["reserve"])

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return items, warnings

    # Trouver la ligne d'en-tête (première ligne avec plusieurs cellules non nulles)
    header_row_idx = None
    header = []
    for i, row in enumerate(all_rows):
        non_null = [c for c in row if c is not None]
        if len(non_null) >= 2:
            header_row_idx = i
            header = [str(c) if c is not None else "" for c in row]
            break

    if header_row_idx is None:
        warnings.append(f"Feuille {sheet_num} : aucune ligne d'en-tête détectée.")
        return items, warnings

    # Détecter les colonnes
    item_col = _find_col(header, _ITEM_SYN)
    qty_col = _find_col(header, _QTY_SYN)
    days_col = _find_col(header, _DAYS_SYN)
    price_col = _find_col(header, _PRICE_SYN)
    total_col = _find_col(header, _TOTAL_SYN)
    supplier_col = _find_col(header, _SUPPLIER_SYN)
    unit_col = _find_col(header, _UNIT_SYN)
    obs_col = _find_col(header, _OBS_SYN)

    if item_col is None:
        warnings.append(
            f"Feuille {sheet_num} : colonne «désignation/item» introuvable dans l'en-tête. "
            f"En-têtes trouvés : {', '.join(c for c in header if c)}."
        )
        return items, warnings

    declared_total_sum = decimal.Decimal(0)
    computed_total_sum = decimal.Decimal(0)
    module_primary = modules[0]

    def get_cell(row: tuple, col_idx: int | None) -> Any:
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    for row in all_rows[header_row_idx + 1:]:
        label_raw = get_cell(row, item_col)
        if label_raw is None or str(label_raw).strip() == "":
            continue
        label = str(label_raw).strip()

        # Ignorer les lignes «Total», «Sous-total», etc.
        if any(kw in label.lower() for kw in ("total", "sous-total", "grand total", "somme")):
            declared_total_raw = get_cell(row, total_col)
            if declared_total_raw is not None:
                v = _to_decimal(declared_total_raw)
                if v is not None:
                    declared_total_sum += v
            continue

        qty = _to_decimal(get_cell(row, qty_col))
        days = _to_decimal(get_cell(row, days_col))
        price = _to_decimal(get_cell(row, price_col))
        declared = _to_decimal(get_cell(row, total_col))
        supplier = str(get_cell(row, supplier_col) or "").strip()
        unit = str(get_cell(row, unit_col) or "").strip()
        obs_val = str(get_cell(row, obs_col) or "").strip().lower()
        is_example = "exemple" in obs_val or "example" in obs_val

        # Calculer le total
        computed: decimal.Decimal | None = None
        if sheet_num == 4 and qty is not None and days is not None and price is not None:
            # Main-d'œuvre : personnes × jours × coût/jour
            computed = qty * days * price
        elif qty is not None and price is not None:
            computed = qty * price
        elif declared is not None:
            computed = declared  # fallback si pas de formule possible

        if declared is not None:
            declared_total_sum += declared
        if computed is not None:
            computed_total_sum += computed

        # Détecter écart déclaré vs calculé
        item_warning = ""
        if declared is not None and computed is not None and abs(computed - declared) > decimal.Decimal("0.01"):
            item_warning = (
                f"Écart total : déclaré {declared}, recalculé {computed:.2f}."
            )

        items.append({
            "module": module_primary,
            "label": label,
            "quantity": str(qty) if qty is not None else None,
            "unit": unit,
            "unit_price": str(price) if price is not None else None,
            "declared_total": str(declared) if declared is not None else None,
            "computed_total": str(computed) if computed is not None else None,
            "suggested_supplier": supplier,
            "_warning": item_warning,
            "_is_example": is_example,
        })

    # Écart total déclaré vs recalculé au niveau de la feuille
    if declared_total_sum > 0 and computed_total_sum > 0:
        ecart_pct = abs(computed_total_sum - declared_total_sum) / declared_total_sum * 100
        if ecart_pct > 5:
            warnings.append(
                f"Feuille {sheet_num} ({module_primary}) : total déclaré {declared_total_sum:.2f} "
                f"≠ total recalculé {computed_total_sum:.2f} "
                f"(écart {ecart_pct:.1f} %)."
            )

    return items, warnings


class NeedsSheetParseError(Exception):
    """Le fichier n'est pas reconnu comme une Feuille de Besoins."""


def parse_needs_sheet(
    file_path: str,
    value_chain=None,
    area_ha: decimal.Decimal | None = None,
    currency: str = "USD",
) -> dict:
    """
    Parse un classeur Excel Feuille de Besoins (feuilles 2-7).
    Retourne un dict {"ok": bool, "items": [...], "totalByModule": {...},
                      "grandTotal": Decimal, "warnings": [...], "anomalies": [...]}.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise NeedsSheetParseError(f"Impossible d'ouvrir le fichier : {exc}") from exc

    # ── Lecture prioritaire depuis la feuille Synthèse ────────────────────────
    synth_totals, synth_grand = _read_synthese_sheet(wb)
    has_synthese = synth_totals is not None and synth_grand is not None and synth_grand > 0

    # ── Validation : la feuille Synthèse OU les sections détail doivent exister ─
    found_sheets = []
    for sheet_num in range(2, 8):
        ws = _find_sheet(wb, sheet_num)
        if ws is not None:
            found_sheets.append(sheet_num)

    if not has_synthese and len(found_sheets) < 3:
        found_names = [ws.title for ws in wb.worksheets]
        raise NeedsSheetParseError(
            f"Le fichier ne ressemble pas à une Feuille de Besoins AGRICAP : "
            f"feuille de synthèse introuvable et seulement {len(found_sheets)} "
            f"section(s) sur 6 reconnues. "
            f"Feuilles trouvées : {', '.join(found_names)}. "
            f"Téléchargez le modèle officiel via GET /api/credits/needs-sheet-template/."
        )

    all_items: list[dict] = []
    all_warnings: list[str] = []

    # ── Parse des lignes de détail (pour fournisseurs et vérifications) ───────
    for sheet_num in range(2, 8):
        ws = _find_sheet(wb, sheet_num)
        if ws is None:
            continue
        items, warnings = _parse_sheet(ws, sheet_num)
        all_items.extend(items)
        all_warnings.extend(warnings)

    wb.close()

    # ── Totaux : Synthèse en priorité, somme des items en fallback ────────────
    if has_synthese:
        total_by_module = synth_totals
        grand_total = synth_grand
    else:
        total_by_module = {}
        for item in all_items:
            mod = item["module"]
            total_by_module[mod] = total_by_module.get(mod, 0) + float(
                item["computed_total"] or item["declared_total"] or 0
            )
        grand_total = sum(total_by_module.values())

    # ── Cohérence vs référentiel ──────────────────────────────────────────────
    anomalies: list[str] = []
    if value_chain is not None and area_ha and area_ha > 0 and grand_total > 0:
        ref_cost = float(value_chain.cost_per_hectare_usd if currency == "USD"
                         else value_chain.cost_per_hectare_cdf) * float(area_ha)
        ratio = grand_total / ref_cost if ref_cost > 0 else 0
        if ratio > 1.30:
            anomalies.append(
                f"Vos besoins ({grand_total:,.0f} {currency}) représentent "
                f"{ratio:.1f}× le coût standard {value_chain.label} "
                f"à {float(area_ha)} ha ({ref_cost:,.0f} {currency}). "
                f"Un justificatif sera demandé à l'analyste."
            )
        elif ratio < 0.70:
            anomalies.append(
                f"Vos besoins ({grand_total:,.0f} {currency}) sont nettement inférieurs "
                f"au coût standard {value_chain.label} "
                f"à {float(area_ha)} ha ({ref_cost:,.0f} {currency}). "
                f"Vérifiez l'exhaustivité du dossier."
            )

        # Cohérence des poids modulaires
        expected_weights = value_chain.module_weights or {}
        for mod, expected_pct in expected_weights.items():
            actual = total_by_module.get(mod, 0)
            if grand_total > 0:
                actual_pct = actual / grand_total * 100
                if expected_pct > 5 and actual_pct > (expected_pct * 1.8):
                    anomalies.append(
                        f"Module «{mod}» : {actual_pct:.0f} % du total "
                        f"(standard : {expected_pct:.0f} %). "
                        f"Vérifiez cette ligne."
                    )

    # ── Détection du gabarit non rempli ──────────────────────────────────────
    example_items = [i for i in all_items if i.get("_is_example")]
    real_items = [i for i in all_items if not i.get("_is_example")]
    if example_items and not real_items:
        raise NeedsSheetParseError(
            "Vous avez téléversé le gabarit vierge AGRICAP sans le modifier. "
            "Veuillez saisir vos besoins réels dans la feuille '4_Besoins_Financiers' "
            "en remplaçant les lignes d'exemple par vos données, puis re-téléversez le fichier."
        )

    # ── Vérification fournisseurs agréés ────────────────────────────────────
    _check_suppliers(all_items, all_warnings)

    # Transférer les warnings item-level vers all_warnings, nettoyer les clés internes
    item_warnings = [i.pop("_warning", "") for i in all_items if i.get("_warning")]
    all_warnings.extend([w for w in item_warnings if w])
    for i in all_items:
        i.pop("_is_example", None)

    return {
        "ok": True,
        "items": all_items,
        "totalByModule": total_by_module,
        "grandTotal": grand_total,
        "warnings": all_warnings,
        "anomalies": anomalies,
    }


def _check_suppliers(items: list[dict], warnings: list[str]) -> None:
    """Vérifie les fournisseurs suggérés contre l'annuaire agréé."""
    try:
        from suppliers.models import Supplier
        known_suppliers = set(
            s.name.lower() for s in Supplier.objects.filter(blacklisted=False)
        )
        blacklisted = set(
            s.name.lower() for s in Supplier.objects.filter(blacklisted=True)
        )
    except Exception:
        return  # suppliers app non disponible

    for item in items:
        supplier = item.get("suggested_supplier", "").strip()
        if not supplier:
            continue
        sl = supplier.lower()
        if sl in blacklisted:
            item["supplier_warning"] = f"Fournisseur «{supplier}» sur liste noire."
            warnings.append(
                f"Fournisseur «{supplier}» (ligne «{item['label'][:50]}») : "
                f"sur liste noire — ce fournisseur ne peut pas être utilisé."
            )
        elif known_suppliers and sl not in known_suppliers:
            item["supplier_warning"] = f"Fournisseur «{supplier}» non agréé."
            warnings.append(
                f"Fournisseur «{supplier}» (ligne «{item['label'][:50]}») : "
                f"non trouvé dans l'annuaire des fournisseurs agréés."
            )
